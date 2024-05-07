import collections
import os
import pandas as pd
import numpy as np
import anndata
import pickle
from matplotlib import pyplot as plt
import matplotlib as mpl
from scipy.stats import gaussian_kde
import matplotlib.gridspec as gridspec
from matplotlib.colors import LinearSegmentedColormap
from lifelines import KaplanMeierFitter, CoxPHFitter
from lifelines.statistics import logrank_test
from lifelines.statistics import multivariate_logrank_test
from scipy import stats
from scipy.stats import linregress, t
from scipy.cluster.hierarchy import linkage, dendrogram, to_tree, fcluster
from ete3 import TreeStyle, Tree, TextFace, NodeStyle, TextFace, AttrFace
import matplotlib.ticker as mtick
from tqdm import tqdm
from PIL import Image
import seaborn as sns
from sklearn.cluster import KMeans
from scipy.ndimage import gaussian_filter
from matplotlib.transforms import Affine2D
from matplotlib.colors import to_rgba
from pyvirtualdisplay import Display
from fig_settings import *
from sklearn.cluster import DBSCAN
from sklearn import metrics
import velocity_plot_utils
from sklearn.preprocessing import LabelEncoder
import scipy.interpolate as interpolate


def load_obj(path):
    with open(path, 'rb') as f:
        return pickle.load(f)
    
def format_ax(ax):
    ax.spines['right'].set_visible(False)
    ax.spines['top'].set_visible(False)
    
def arrow_axes(ax, xlabel='x', ylabel='y', l=0.3, x_shift=-0.04, y_shift=-0.03):
    ax.set_xticks([])
    ax.set_yticks([])
    # Remove spines
    ax.spines[['top', 'right', 'bottom', 'left']].set_visible(False)
    clr = ax.spines['bottom'].get_edgecolor()
    # Add arrows
    ax.arrow(0, 0, l, 0, transform=ax.transAxes, fc=clr, ec=clr, \
            lw=1, head_width=0.03, head_length=0.03, \
            length_includes_head=True, clip_on=False)
    ax.arrow(0, 0, 0, l, transform=ax.transAxes, fc=clr, ec=clr, \
            lw=1, head_width=0.03, head_length=0.03, \
            length_includes_head=True, clip_on=False)
    ax.text(l/2, 0 + x_shift, xlabel, transform=ax.transAxes, \
        verticalalignment='center', horizontalalignment='center')
    ax.text(0 + y_shift, l/2, ylabel, transform=ax.transAxes, rotation=90,\
        verticalalignment='center', horizontalalignment='center')

def load_df_list(source_data_path, wb_name, div_rows, div_cols, n_div):
	'''
	source_data_path: path to the source data
	wb_name: name of the worksheet
	div_rows: number of rows in each division
	div_cols: number of columns in each division
	n_div: number of divisions
	'''
	import openpyxl
	wb = openpyxl.load_workbook(source_data_path, data_only=True)
	sheet = wb[wb_name]
	df_list = collections.defaultdict(list)
	for i in range(n_div):
		metric = sheet.cell(row=i*(div_rows+1) + 1, column=1).value
		div_names, div_values = [], []
		for j in range(2, div_rows+1):
			m_name = sheet.cell(row=i*(div_rows+1) + j, column=1).value
			vals = [sheet.cell(row=i*(div_rows+1) + j, column=k).value for k in range(2, div_cols + 1)]
			div_settings = [sheet.cell(row=i*(div_rows+1) + 1, column=k).value for k in range(2, div_cols + 1)]
			div_names.append(m_name)
			div_values.append(vals)
		div_values = np.array(div_values)
		if metric == 'F1':
			metric = r'$F_1$'
		df_list[metric] = pd.DataFrame(div_values, index=div_names, columns=div_settings)
	return df_list
    
def load_gdsc_data(source_data_path='.xlsx'):
	"""
	Load GDSC data
	:param source_data_path: path to the source data
	:return: a dictionary of dataframes
	: dataframe: index: methods, columns: settings, values: evaluation metrics
	"""
	div_rows, div_cols, n_div =7, 16, 4
	df_list = load_df_list(source_data_path, 'Drug Combo GDSC', div_rows, div_cols, n_div)
	return df_list

def load_eval_topk_data(source_data_path='.xlsx'):
    """
    Load GDSC data
    :param source_data_path: path to the source data
    :return: a dictionary of dataframes
    : dataframe: index: methods, columns: settings, values: evaluation metrics
    """
    div_rows, div_cols, n_div =7, 6, 4
    df_list = load_df_list(source_data_path, 'Eval top k', div_rows, div_cols, n_div)
    return df_list

def load_drugbank_data(source_data_path):
	div_rows, div_cols, n_div =8, 7, 3
	df_list = load_df_list(source_data_path, 'DDI DrugBank', div_rows, div_cols, n_div)
	return df_list

def load_twosides_data(source_data_path):
    div_rows, div_cols, n_div =7, 3, 3
    df_list = load_df_list(source_data_path, 'DDI Two-sides', div_rows, div_cols, n_div)
    return df_list

def load_xenograft_response(source_data_path):
	div_rows, div_cols, n_div =6, 11, 2
	df_list = load_df_list(source_data_path, 'Xenograft', div_rows, div_cols, n_div)
	return df_list

def load_gdsc_3_drug(source_data_path):
	import openpyxl
	wb = openpyxl.load_workbook(source_data_path, data_only=True)
	sheet = wb['GDSC 3 drug Combo per cell line']
	cell_line, auroc, auprc = [], [], []
	n_rows = sheet.max_row
	for i in range(2, n_rows):
		cell_line.append(sheet.cell(row=i, column=1).value)
		auroc.append(sheet.cell(row=i, column=2).value)
		auprc.append(sheet.cell(row=i, column=3).value)
	return cell_line, auroc, auprc

def load_gdsc_3_drug_preds(source_data_path):
    import openpyxl
    wb = openpyxl.load_workbook(source_data_path, data_only=True)
    sheet = wb['GDSC 3 drug Combo all preds']
    cmb2preds = collections.OrderedDict()
    n_rows = sheet.max_row
    for i in range(2, n_rows):
        d_a, d_b, d_c, cell_line = sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=2).value, \
                                      sheet.cell(row=i, column=3).value, sheet.cell(row=i, column=4).value
        pred = sheet.cell(row=i, column=5).value
        target = sheet.cell(row=i, column=6).value
        cmb = '+'.join(sorted([d_a, d_b, d_c]))
        if cmb not in cmb2preds:
            cmb2preds[cmb] = collections.OrderedDict()
            cmb2preds[cmb]['cell_line'] = []
            cmb2preds[cmb]['pred'] = []
            cmb2preds[cmb]['target'] = []
        cmb2preds[cmb]['cell_line'].append(cell_line)
        cmb2preds[cmb]['pred'].append(pred)
        cmb2preds[cmb]['target'].append(target)
    return cmb2preds

def load_cell2cancer():
    this_file = os.path.abspath(__file__)
    cell_cancer_dir = os.path.join(os.path.dirname(this_file), '../Pisces/scripts/case_study/gdsc_novel_combo/data/cell_cancer_info.csv')
    cell_cancer_df = pd.read_csv(cell_cancer_dir)
    cell2cancer = collections.OrderedDict()
    for i in range(cell_cancer_df.shape[0]):
        cell = cell_cancer_df.iloc[i, 1]
        cancer = cell_cancer_df.iloc[i, 0]
        cell2cancer[cell] = cancer
    return cell2cancer

def load_best_response_preds(source_data_path):
    import openpyxl
    wb = openpyxl.load_workbook(source_data_path, data_only=True)
    sheet = wb['Xenograft']
    n_rows = sheet.max_row
    is_data = 0
    preds, labels = [], []
    for i in range(1, n_rows):
        if sheet.cell(row=i, column=1).value == 'combo name' and sheet.cell(row=i-1, column=1).value == 'Best Response':
            is_data = 1
            continue
        if is_data == 1:
            if sheet.cell(row=i, column=1).value is None:
                print('empty cell at row {}'.format(i))
                break
            preds.append(sheet.cell(row=i, column=2).value)
            labels.append(sheet.cell(row=i, column=3).value)
    return np.asarray(preds), np.asarray(labels)

def load_days_response_preds(source_data_path):
    import openpyxl
    wb = openpyxl.load_workbook(source_data_path, data_only=True)
    sheet = wb['Xenograft']
    n_rows = sheet.max_row
    is_data = 0
    preds, labels = [], []
    for i in range(1, n_rows):
        if sheet.cell(row=i, column=5).value == 'combo name' and sheet.cell(row=i-1, column=5).value == 'Days Response':
            is_data = 1
            continue
        if is_data == 1:
            if sheet.cell(row=i, column=6).value is None:
                print('empty cell at row {}'.format(i))
                break
            preds.append(sheet.cell(row=i, column=6).value)
            labels.append(sheet.cell(row=i, column=7).value)
    return np.asarray(preds), np.asarray(labels)

def load_days_response_extrapolation_preds(source_data_path):
    import openpyxl
    wb = openpyxl.load_workbook(source_data_path, data_only=True)
    sheet = wb['Xenograft']
    n_rows = sheet.max_row
    is_data = 0
    preds, labels = [], []
    for i in range(1, n_rows):
        if sheet.cell(row=i, column=9).value == 'combo name' and sheet.cell(row=i-1, column=9).value == 'Days Response Extrapolation':
            is_data = 1
            continue
        if is_data == 1:
            if sheet.cell(row=i, column=10).value is None:
                print('empty cell at row {}'.format(i))
                break
            preds.append(sheet.cell(row=i, column=10).value)
            labels.append(sheet.cell(row=i, column=11).value)
    return np.asarray(preds), np.asarray(labels)

def load_days_response_umap(source_data_path, fold=0):
    import openpyxl
    wb = openpyxl.load_workbook(source_data_path, data_only=True)
    sheet = wb['Xenograft']
    n_rows = sheet.max_row
    is_data = 0
    names = []
    umaps_array = []
    preds, labels = [], []
    for i in range(1, n_rows):
        if sheet.cell(row=i, column=13).value == 'combo name' and sheet.cell(row=i-1, column=13).value == 'Days Response UMAP':
            is_data = 1
            continue
        if is_data == 1:
            if sheet.cell(row=i, column=14).value is None:
                print('empty cell at row {}'.format(i))
                break
            if 'fold{}'.format(fold) not in sheet.cell(row=i, column=13).value:
                continue
            names.append(sheet.cell(row=i, column=13).value)
            umaps_array.append([sheet.cell(row=i, column=14).value, sheet.cell(row=i, column=15).value])
            preds.append(sheet.cell(row=i, column=16).value)
            labels.append(sheet.cell(row=i, column=17).value)
    return np.asarray(names), np.asarray(umaps_array), np.asarray(preds), np.asarray(labels)

def load_days_response_umap_from_csv(csv_data_path, fold=0):
    df = pd.read_csv(csv_data_path)
    n_rows = df.shape[0]
    names = []
    umaps_array = []
    preds, labels = [], []
    for i in range(n_rows):
        if 'fold{}'.format(fold) not in df.iloc[i, 0]:
            continue
        names.append(df.iloc[i, 0])
        umaps_array.append([df.iloc[i, 1], df.iloc[i, 2]])
        preds.append(df.iloc[i, 3])
        labels.append(df.iloc[i, 4])
    return np.asarray(names), np.asarray(umaps_array), np.asarray(preds), np.asarray(labels)

def load_gdsc_trans_eval_modalities(source_data_path):
    import openpyxl
    wb = openpyxl.load_workbook(source_data_path, data_only=True)
    sheet = wb['GDSC Evaluate Modalities']
    n_rows = sheet.max_row
    modality2bacc, modality2auprc = collections.OrderedDict(), collections.OrderedDict()
    modality2f1, modality2kappa = collections.OrderedDict(), collections.OrderedDict()
    for i in range(1, n_rows):
        if sheet.cell(row=i, column=1).value == 'fold':
            continue
        modality = sheet.cell(row=i, column=2).value
        if modality not in modality2bacc:
            modality2bacc[modality] = []
            modality2auprc[modality] = []
            modality2f1[modality] = []
            modality2kappa[modality] = []
        modality2bacc[modality].append(sheet.cell(row=i, column=3).value)
        modality2auprc[modality].append(sheet.cell(row=i, column=4).value)
        modality2f1[modality].append(sheet.cell(row=i, column=5).value)
        modality2kappa[modality].append(sheet.cell(row=i, column=6).value)
    return modality2bacc, modality2auprc, modality2f1, modality2kappa

def select_cell_line(cell_line, auroc, auprc, cell2cancer, cancer):
	sel_cell_line = []
	sel_auroc, sel_auprc = [], []
	for idx, c in enumerate(cell_line):
		if c not in cell2cancer:
			continue
		if cell2cancer[c] == cancer:
			sel_cell_line.append(c)
			sel_auroc.append(auroc[idx])
			sel_auprc.append(auprc[idx])
	return sel_cell_line, sel_auroc, sel_auprc

def circular_bar_plot(ax, names, values, mean_value, y_labels, hie_colors):
    '''
    The function is used to plot circular bar plot
	:param ax: axis
	:param names: names of the methods
	:param values: values of the methods
	:param mean_value: mean value
	:param y_labels: y labels
    '''
    width = 0.1
    bottom_start = 1
    srt_names = list(names)
    srt_values = list(values)
    angles = np.linspace(np.pi, -np.pi, len(values) + 4, endpoint=False)
    hie_colors.append(hie_colors[0])

    srt_values.append(mean_value)
    srt_names.append('Mean')
    ct_angles = angles[0: len(srt_values)]
    ax.bar(ct_angles, np.asarray(srt_values), width=width, 
			color=hie_colors, edgecolor='k', 
			bottom=bottom_start, zorder=3)
    for angle, label, val in zip(ct_angles, srt_names, srt_values):
        ax.text(angle, bottom_start + val + 0.3, label, 
			rotation=np.rad2deg(angle) + 180,
			ha='center', va='center')
        # set the colors of that bar

    ax.set_xticklabels([])
    ax.set_yticklabels([])
    rgrids = [0, 0.2, 0.4, 0.6, 0.8, 1.0]
    rgrids = [bottom_start + r for r in rgrids]
    ax.set_rgrids(rgrids, color='black')
    ax.set_thetagrids([], [])
    ax.grid(True, color=ax.spines['polar'].get_edgecolor(), linestyle='--', alpha=0.5)
    y_label_angle = np.pi * (len(srt_values) + 7) / (len(srt_values) + 3)
    for y_l in y_labels:
        ax.text(y_label_angle, bottom_start + float(y_l[6:]) - 0.15, y_l,
                rotation=np.rad2deg(y_label_angle) - 90,
                fontsize=SMALLER_SIZE,
                ha='center', va='center')
    ax.spines['polar'].set_visible(False)
    return srt_names, srt_values

def get_cell_line_emb(cell_lines):
    this_file_dir = os.path.dirname(os.path.abspath(__file__))
    emb_path = os.path.join(this_file_dir, '../data/drug_combo/cell_tpm.csv')
    cell2emb = collections.OrderedDict()
    emb_csv = pd.read_csv(emb_path, index_col=0)
    cols = emb_csv.columns
    genes = list(cols[1:])
    for i in range(emb_csv.shape[0]):
        cell = emb_csv.iloc[i]['cell_line_names']
        gene_emb = np.asarray(emb_csv.iloc[i][genes])
        if cell not in cell_lines:
            continue
        cell2emb[cell] = gene_emb
    cell_embs = [cell2emb[c] for c in cell_lines]
    return np.asarray(cell_embs, dtype=np.float32)


def get_newick(node, newick, parentdist, leaf_names):
    if node.is_leaf():
        return f"{newick}{leaf_names[node.id]}:{parentdist - node.dist:.2f}"
    else:
        if len(newick) > 0:
            newick = f"{newick}({get_newick(node.left, '', node.dist, leaf_names)},{get_newick(node.right, '', node.dist, leaf_names)}):{parentdist - node.dist:.2f}"
        else:
            newick = f"({get_newick(node.left, '', node.dist, leaf_names)},{get_newick(node.right, '', node.dist, leaf_names)}):{parentdist - node.dist:.2f}"
        return newick

def circular_layout(node):
    l_width=30
    node_style = NodeStyle()
    node_style["size"] = 50
    node_style["hz_line_width"] = l_width  # Adjust horizontal line width
    node_style["vt_line_width"] = l_width
    node_style["hz_line_color"] = "#f0f0f0"
    node_style["vt_line_color"] = "#f0f0f0"
    node.set_style(node_style)

def hiearchical_clustering(ax, embs, names, s, max_c=12):
    display = Display(visible=0, size=(s, s))
    display.start()
    emb_log2 = np.log2(embs + 1)
    print('start clustering')
    linkage_matrix = linkage(emb_log2, method='average', metric='euclidean')
    cluster_labels = fcluster(linkage_matrix, max_c, criterion='maxclust')
    cluster_colors = {name: get_cluster_colors(label) for name, label in zip(names, cluster_labels)}
    tree = to_tree(linkage_matrix)
    leaf_names = list(names)
    newick = get_newick(tree, "", tree.dist, leaf_names) + ";"

	# Create a circular tree plot using ete3
    t = Tree(newick)
    idx = 0
    srt_names, colors = [], []
    for leaf in t.iter_leaves():
        srt_names.append(leaf.name)
        colors.append(cluster_colors[leaf.name])
        idx += 1
    # adjust the top 3 leaf nodes
    '''idx = 0 
    for leaf in t.iter_leaves():
        if idx < 15:
            if idx < 9: 
                leaf.dist *= 1.3
            else: 
                leaf.dist *= 1.2
        idx += 1'''
    ts = TreeStyle()
    ts.mode = "c"  # Circular mode
    ts.arc_start = 180  # 0 degrees = 3 o'clock
    ts.arc_span = 360 * (len(names)) / (len(names) + 4) - 2
    ts.show_scale = False
    ts.aligned_header = True
    ts.show_leaf_name = False
    ts.layout_fn = circular_layout
	# Adjust the scale property to control the size of the tree
    ts.scale = 10  # Adjust the scale property to control the size of the tree
    ts.margin_left = 1
    ts.margin_right = 1
    ts.margin_top = 1
    ts.margin_bottom = 1
	# Display the circular tree plot
    this_file_dir = os.path.dirname(os.path.abspath(__file__))
    save_path = os.path.join(this_file_dir, 'save_figs/circular_tree.png')
    t.render(save_path, w=800, units='px', tree_style=ts)
    display.stop()
    return srt_names, colors

def gdsc_3_drug_auroc_barplot(cell_line, auroc, fig_suffix='.png'):
    fig, ax = get_circular_plot_axis()
    fig.subplots_adjust(left=0.1, bottom=0.1, right=0.9, top=0.9, wspace=None, hspace=None)

    names, values = cell_line[:-1], auroc[:-1] 
    mean_value = auroc[-1]
    y_labels = ['AUROC=0.2', 'AUROC=0.4', 'AUROC=0.6', 
                        'AUROC=0.8', 'AUROC=1.0']
    
    cell_emb = get_cell_line_emb(names)
    hie_names, hie_colors = hiearchical_clustering(ax, cell_emb, names, s=800)
    print(hie_names)
    this_file_dir = os.path.dirname(os.path.abspath(__file__))
    img = Image.open(os.path.join(this_file_dir, 'save_figs/circular_tree.png')).copy()
    tree_image_array = np.array(img.convert('RGBA'))
    white_color = (255, 255, 255)
    tree_image_array[np.all(tree_image_array[:, :, :3] == white_color, axis=2), 3] = 0
    #ax_image = fig.add_axes([0.31, 0.31, 0.38, 0.38])
    ax_image = fig.add_axes([0.3, 0.3, 0.4, 0.4])
    ax_image.imshow(tree_image_array ,zorder=0)
    ax_image.axis('off')
    
    # sort the names and values using hie_names
    srt_idx = [names.index(n) for n in hie_names]
    srt_names = list(np.array(names)[srt_idx])
    srt_values = list(np.array(values)[srt_idx])
    circular_bar_plot(ax, srt_names, srt_values, mean_value, y_labels, hie_colors)

    this_file_dir = os.path.dirname(os.path.abspath(__file__))
    fig.savefig(os.path.join(this_file_dir, f'save_figs/gdsc_3_drug_cell_line_auroc{fig_suffix}'), dpi=300)
    
def gdsc_3_drug_auprc_barplot(cell_line, auprc, fig_suffix='.png'):
    fig, ax = get_circular_plot_axis()
    fig.subplots_adjust(left=0.1, bottom=0.1, right=0.9, top=0.9, wspace=None, hspace=None)

    names, values = cell_line[:-1], auprc[:-1] 
    mean_value = auprc[-1]
    y_labels = ['AUPRC=0.2', 'AUPRC=0.4', 'AUPRC=0.6', 
                        'AUPRC=0.8', 'AUPRC=1.0']
    
    cell_emb = get_cell_line_emb(names)
    hie_names, hie_colors = hiearchical_clustering(ax, cell_emb, names, s=800)
    print(hie_names)
    this_file_dir = os.path.dirname(os.path.abspath(__file__))
    img = Image.open(os.path.join(this_file_dir, 'save_figs/circular_tree.png')).copy()
    tree_image_array = np.array(img.convert('RGBA'))
    white_color = (255, 255, 255)
    tree_image_array[np.all(tree_image_array[:, :, :3] == white_color, axis=2), 3] = 0
    #ax_image = fig.add_axes([0.31, 0.31, 0.38, 0.38])
    ax_image = fig.add_axes([0.3, 0.3, 0.4, 0.4])
    ax_image.imshow(tree_image_array ,zorder=0)
    ax_image.axis('off')
    
    # sort the names and values using hie_names
    srt_idx = [names.index(n) for n in hie_names]
    srt_names = list(np.array(names)[srt_idx])
    srt_values = list(np.array(values)[srt_idx])
    circular_bar_plot(ax, srt_names, srt_values, mean_value, y_labels, hie_colors)

    this_file_dir = os.path.dirname(os.path.abspath(__file__))
    fig.savefig(os.path.join(this_file_dir, f'save_figs/gdsc_3_drug_cell_line_auprc{fig_suffix}'), dpi=300)

def gdsc_3_drug_barplot(cell_line, auroc, auprc, fig_suffix='.png'):
    gdsc_3_drug_auroc_barplot(cell_line, auroc, fig_suffix=fig_suffix)
    gdsc_3_drug_auprc_barplot(cell_line, auprc, fig_suffix=fig_suffix)

def draw_gradient_square(x, y, size, color1, color2, ax):
    gradient = np.array([np.linspace(c1/255.0, c2/255.0, size) for c1, c2 in zip(color1, color2)])
    square_gradient = np.stack([gradient] * size, axis=1)
    extent = [x, x + 1, y, y + 1]
    im = ax.imshow(square_gradient.transpose(2,1,0), extent=extent, origin='lower', aspect='auto')

def gdsc_3_drug_waterfall_plot(cmb2preds, fig_suffix='.png'):
    from sklearn.metrics import roc_auc_score, average_precision_score
    fig, ax = get_waterfall_axis()
    gs = gridspec.GridSpec(2, 1, height_ratios=[2, 3])
    ax_waterfall = plt.subplot(gs[1, 0])
    ax_bar = plt.subplot(gs[0, 0], sharex=ax_waterfall)
    
    cmb_names, pred_list, target_list, pred_mean = [], [], [], []
    auroc_list, auprc_list = [], []
    min_auroc, min_auprc = 1, 1
    for cmb, preds in cmb2preds.items():
        pred, tgt = preds['pred'], preds['target']
        srt_idx = np.argsort(pred)
        pred = [pred[i] for i in srt_idx]
        tgt = [tgt[i] for i in srt_idx]
        cmb_names.append(cmb)
        pred_list.append(pred)
        pred_mean.append(np.mean(preds['pred']))
        if np.max(tgt) == 0:
            auroc = 'Null'
            auprc = 'Null'
        else:
            auroc = roc_auc_score(tgt, pred)
            auprc = average_precision_score(tgt, pred)
            if auroc < min_auroc:
                min_auroc = auroc
            if auprc < min_auprc:
                min_auprc = auprc
        auroc_list.append(auroc)
        auprc_list.append(auprc)
    print('min auroc', min_auroc)
    print('min auprc', min_auprc)
    valid_auroc_list = [a for a in auroc_list if a != 'Null']
    valid_auprc_list = [a for a in auprc_list if a != 'Null']
    print('mean auroc', np.mean(valid_auroc_list))
    print('mean auprc', np.mean(valid_auprc_list))
    pred_mean = np.array(pred_mean)
    srt_idx = np.argsort(pred_mean)
    cmb_names = [cmb_names[i] for i in srt_idx]
    pred_list = [pred_list[i] for i in srt_idx]
    auroc_list = [auroc_list[i] for i in srt_idx]
    auprc_list = [auprc_list[i] for i in srt_idx]
    pred_mean = pred_mean[srt_idx]
    for i in range(len(cmb_names)):
        cmb = cmb_names[i]
        preds = pred_list[i]
        x = np.arange(len(preds))
        x = 0.8 * x / len(preds) + 0.1 + i
        ax_waterfall.scatter(x, preds, linewidth=1, s=5)
        ax_waterfall.plot([i, i], [0, 1], color='w', linewidth=0.5)
        ax_waterfall.plot([i+1, i+1], [0, 1], color='w', linewidth=0.5)
        ax_waterfall.plot([i, i+1], [pred_mean[i], pred_mean[i]], color='w', linewidth=1)
        ax_waterfall.scatter([i+0.5], [pred_mean[i]], color='w', s=10)
    print('most effective combinations{}, synergy probability{}'.format(cmb_names[-1], pred_mean[-1]))
    ax_waterfall.set_ylim([0, 1])
    ax_waterfall.set_xlim([0, len(cmb_names)])
    ax_waterfall.set_xticks(np.arange(len(cmb_names)) + 0.5)
    xt_labels = []
    for cmb in cmb_names:
        d = cmb.split('+')
        d.remove('Trametinib')
        d.remove('Afatinib')
        xt_labels.append('+ {}'.format(d[0]))
    ax_waterfall.set_xticklabels(xt_labels, rotation=90)
    ax_waterfall.set_ylabel('Synergy (probability)')
    bar_width=0.3
    label_flag = 0
    edgecolor = 'w'
    linewidth = 0.5
    for i in range(len(cmb_names)):
        if auroc_list[i] == 'Null':
            draw_gradient_square(i, 0, 300, (115,115,115,255), (115,115,115,255), ax_bar)
            continue
        if label_flag == 0:
            ax_bar.bar(i+0.5-bar_width/2, auroc_list[i], color='#80cdc1', width=bar_width, edgecolor=edgecolor, linewidth=linewidth, label='AUROC')
            ax_bar.bar(i+0.5+bar_width/2, auprc_list[i], color='#f6e8c3', width=bar_width, edgecolor=edgecolor, linewidth=linewidth, label='AUPRC')
            ax_bar.bar(i+0.5+bar_width/2, auprc_list[i], color='#737373', alpha=1, width=0*bar_width, label='Not available')
            label_flag = 1
        else:
            ax_bar.bar(i+0.5-bar_width/2, auroc_list[i], color='#80cdc1', width=bar_width, edgecolor=edgecolor, linewidth=linewidth)
            ax_bar.bar(i+0.5+bar_width/2, auprc_list[i], color='#f6e8c3', width=bar_width, edgecolor=edgecolor, linewidth=linewidth)

    #ax_bar.set_xticks([])
    format_ax(ax_bar)
    ax_bar.set_ylabel('AUROC/AUPRC')
    ax_bar.legend(loc='upper center', bbox_to_anchor=(0.5, 1.3), ncol=3, frameon=False)
    plt.setp(ax_bar.get_xticklabels(), visible=False)
    fig.tight_layout()
    this_file_dir = os.path.dirname(os.path.abspath(__file__))
    fig.savefig(os.path.join(this_file_dir, f'save_figs/gdsc_3_drug_waterfall_plot{fig_suffix}'), dpi=300)

def xenograt_pearson_barplot(df, fig_suffix='.png'):
	yticks = [[0.2, 0.4, 0.6, 0.8], ['0.2', '0.4', '0.6', '0.8']]
	methods = list(df.index)
	values = np.asarray(df)
	mean_values = np.mean(values, axis=1)
	std_values = np.std(values, axis=1)
	fig, ax = get_narrow_axis()
	# plot
	width = 1
	for i in range(len(methods)):
		ax.bar(i, mean_values[i], \
				width, color=get_model_colors(methods[i]), label=methods[i], \
				edgecolor='k', linewidth=get_single_group_bar_plot_setting()['linewidth'])
		ax.errorbar(i, mean_values[i], \
					yerr=std_values[i], color='black', capsize=get_single_group_bar_plot_setting()['capsize'], \
					capthick=get_single_group_bar_plot_setting()['capthick'], fmt='none', elinewidth=1)
	dots_x, dots_y = [], []
	for i in range(len(methods)):
		for j in range(3):
			dots_x.append(i)
			dots_y.append(values[i, j])
	add_bar_plot_dots(ax, dots_x, dots_y, jitter=width*0.1)

	ax.spines['right'].set_visible(False)
	ax.spines['top'].set_visible(False)
	ax.set_xticklabels(methods)
	ax.set_xticks(np.arange(len(methods)))
	ax.set_ylabel('Pearson correlation')
	ax.set_yticks(yticks[0])
	ax.set_yticklabels(yticks[1])
	plt.setp(ax.get_xticklabels(), rotation=45, ha='right',
             rotation_mode="anchor")
	fig.tight_layout()
	this_file_dir = os.path.dirname(os.path.abspath(__file__))
	fig.savefig(os.path.join(this_file_dir, f'save_figs/xenograft_pearson_barplot{fig_suffix}'), dpi=300)

def box_plot(ax, values, methods, colors):
	boxplots = ax.boxplot(values.T, patch_artist = True, showfliers=False)
	for patch, color in zip(boxplots['boxes'], colors):
		rgba_color = to_rgba(color, alpha=get_box_plot_setting()['box alpha'])
		patch.set_facecolor(rgba_color)
		patch.set_linewidth(get_box_plot_setting()['box linewidth'])
	for median in boxplots['medians']:
		median.set(color=get_box_plot_setting()['median color'],
				linewidth=get_box_plot_setting()['median linewidth'])
	for whisker in boxplots['whiskers']:
		whisker.set(linewidth=get_box_plot_setting()['whisker linewidth'],
					linestyle=get_box_plot_setting()['whisker linestyle'])
	for cap in boxplots['caps']:
		cap.set(linewidth=get_box_plot_setting()['cap linewidth'])
	for i in range(len(methods)):
		dots_x, dots_y = [], []
		for j in range(10):
			dots_x.append(i + 1)
			dots_y.append(values[i, j])
		add_box_plot_dots(ax, dots_x, dots_y, jitter=0.12, color=colors[i])
	ax.spines['right'].set_visible(False)
	ax.spines['top'].set_visible(False)

def xenograt_spearman_boxplot(df, fig_suffix='.png'):
	yticks = [[0.0, 0.2, 0.4, 0.6, 0.8, 1.0], ['0.0', '0.2', '0.4', '0.6', '0.8', '1.0']]
	methods = list(df.index)
	values = np.asarray(df)
	srt_idx = np.argsort(np.median(values, axis=1))[::-1]
	methods = np.array(methods)[srt_idx]
	values = values[srt_idx]
	colors = [get_model_colors(m) for m in methods]
	fig, ax = get_square_axis()
	# Create box plot
	box_plot(ax, values, methods, colors)
	# sort the df using the current methods
	srt_df = df.loc[methods]
	add_significance_symbol_boxplot(ax, srt_df, our_method_index=0)
	# Customize the plot
	ax.set_xticklabels(methods)
	ax.set_yticks(yticks[0])
	ax.set_yticklabels(yticks[1])
	ax.set_ylabel('Spearman correlation')
	# Save the plot as an image
	plt.setp(ax.get_xticklabels(), rotation=45, ha='right',
             rotation_mode="anchor")
	fig.tight_layout()
	this_file_dir = os.path.dirname(os.path.abspath(__file__))
	fig.savefig(os.path.join(this_file_dir, f'save_figs/xenograft_spearman_boxplot{fig_suffix}'), dpi=300)
 
def xenograt_pearson_boxplot(df, fig_suffix='.png'):
	yticks = [[0.0, 0.2, 0.4, 0.6, 0.8], ['0.0', '0.2', '0.4', '0.6', '0.8']]
	methods = list(df.index)
	values = np.asarray(df)
	srt_idx = np.argsort(np.mean(values, axis=1))[::-1]
	methods = np.array(methods)[srt_idx]
	values = values[srt_idx]
	colors = [get_model_colors(m) for m in methods]
	fig, ax = get_square_axis()
	# Create box plot
	box_plot(ax, values, methods, colors)
	# sort the df using the current methods
	srt_df = df.loc[methods]
	add_significance_symbol_boxplot(ax, srt_df, our_method_index=0)
	# Customize the plot
	ax.set_xticklabels(methods)
	ax.set_yticks(yticks[0])
	ax.set_yticklabels(yticks[1])
	ax.set_ylabel('Pearson correlation')
	# Save the plot as an image
	plt.setp(ax.get_xticklabels(), rotation=45, ha='right',
             rotation_mode="anchor")
	fig.tight_layout()
	this_file_dir = os.path.dirname(os.path.abspath(__file__))
	fig.savefig(os.path.join(this_file_dir, f'save_figs/xenograft_pearson_boxplot{fig_suffix}'), dpi=300)

def gdsc_bar_plot(gdsc_list, fig_suffix='.png', n_fold=5):
	yticks =[[[0.45, 0.5, 0.55, 0.6, 0.65, 0.7, 0.75], ['0.45', '0.50', '0.55', '0.60', '0.65', '0.70', '0.75']],
	   		 [[0.0, 0.1, 0.2, 0.3, 0.4, 0.5], ['0.0', '0.1', '0.2', '0.3', '0.4', '0.5']],
			 [[0.0, 0.1, 0.2, 0.3, 0.4, 0.5], ['0.0', '0.1', '0.2', '0.3', '0.4', '0.5']],
			 [[0.0, 0.1, 0.2, 0.3, 0.4, 0.5], ['0.0', '0.1', '0.2', '0.3', '0.4', '0.5']]]

	fig, axes = get_four_square_axis()
	x_ticklabels = ['Vanilla \ncross validation', 'Split by \ncombination', 'Split by \ncell line']
	for idx, (metric, df) in enumerate(gdsc_list.items()):
		bar_plot_exclude_graphsynergy(df, axes[idx], n_group=3, n_fold=n_fold, \
                edgecolor=axes[idx].spines['bottom'].get_edgecolor(), \
	   			xtick_label=x_ticklabels, capsize=1.5, capthick=1)
		axes[idx].set_ylabel(metric)
		axes[idx].set_yticks(yticks[idx][0])
		axes[idx].set_yticklabels(yticks[idx][1])
		plt.setp(axes[idx].get_xticklabels(), rotation=45, ha='right',
             rotation_mode="anchor")
		axes[idx].set_ylim(yticks[idx][0][0], yticks[idx][0][-1])
		add_significance_symbol_barplot(axes[idx], df, our_method_index=0, n_group=3, n_fold=n_fold)
 
	handles, labels = axes[idx].get_legend_handles_labels()
	fig.legend(handles, labels, loc='upper center', 
            #bbox_to_anchor=(0.5, 1.05),
            ncol=6, frameon=False)
	for idx in range(4):
		x0, x1 = axes[idx].get_xlim()
		y0, y1 = axes[idx].get_ylim()
		axes[idx].set_aspect(1.45*abs(x1-x0)/abs(y1-y0))
	fig.tight_layout()
	this_file_dir = os.path.dirname(os.path.abspath(__file__))
	fig.savefig(os.path.join(this_file_dir, f'save_figs/gdsc_barplot{fig_suffix}'), dpi=300)
 

def drugbank_bar_plot(drugbank_list, fig_suffix='.png'):
	yticks =[[[0.60, 0.65, 0.70, 0.75, 0.80, 0.85, 0.90, 0.95, 1.00], ['0.60', '0.65', '0.70', '0.75', '0.80', '0.85', '0.90', '0.95', '1.00']],
			 [[0.70, 0.75, 0.80, 0.85, 0.90, 0.95, 1.00], ['0.70', '0.75', '0.80', '0.85', '0.90', '0.95', '1.00']],
			 [[0.60, 0.65, 0.70, 0.75, 0.80, 0.85, 0.90, 0.95, 1.00], ['0.60', '0.65', '0.70', '0.75', '0.80', '0.85', '0.90', '0.95', '1.00']],]

	fig, axes = get_three_square_axis()
	x_ticklabels = ['Vanilla cross validation', 'One new drug \n in each test pair', 'Two new drugs \n in each test pair']
	for idx, (metric, df) in enumerate(drugbank_list.items()):
		bar_plot_mean_std(df, axes[idx], n_group=3, \
                    edgecolor=axes[idx].spines['bottom'].get_edgecolor(), \
					xtick_label=x_ticklabels, capsize=1.5, capthick=1)
		axes[idx].set_ylabel(metric)
		axes[idx].set_yticks(yticks[idx][0])
		axes[idx].set_yticklabels(yticks[idx][1])
		plt.setp(axes[idx].get_xticklabels(), rotation=45, ha='right',
				rotation_mode="anchor")
		axes[idx].set_ylim(yticks[idx][0][0], yticks[idx][0][-1])
		add_significance_symbol_barplot_mean_std(axes[idx], df, our_method_index=0, n_group=3, n_fold=3)

	handles, labels = axes[idx].get_legend_handles_labels()
	fig.legend(handles, labels, loc='upper center', ncol=7, frameon=False)
	for idx in range(len(drugbank_list)):
		x0, x1 = axes[idx].get_xlim()
		y0, y1 = axes[idx].get_ylim()
		axes[idx].set_aspect(abs(x1-x0)/abs(y1-y0))
	fig.tight_layout()
	this_file_dir = os.path.dirname(os.path.abspath(__file__))
	fig.savefig(os.path.join(this_file_dir, f'save_figs/drugbank_barplot{fig_suffix}'), dpi=300)


def two_sides_bar_plot(two_sides_list, fig_suffix='.png'):
	yticks =[[[0.40, 0.50, 0.60, 0.70, 0.80, 0.90], ['0.4', '0.5', '0.6', '0.7', '0.8', '0.9']],
			 [[0.40, 0.50, 0.60, 0.70, 0.80, 0.90], ['0.4', '0.5', '0.6', '0.7', '0.8', '0.9']],
             [[0.40, 0.50, 0.60, 0.70, 0.80, 0.90, 1.0], ['0.4', '0.5', '0.6', '0.7', '0.8', '0.9', '1.0']],]
	fig, axes = get_three_narrow_axis()
	for idx, (metric, df) in enumerate(two_sides_list.items()):
		bar_plot_mean_std_per_method(df, axes[idx], 
                            edgecolor=axes[idx].spines['bottom'].get_edgecolor(),
                            capsize=get_single_group_bar_plot_setting()['capsize'], capthick=get_single_group_bar_plot_setting()['capthick'], 
                            linewidth=get_single_group_bar_plot_setting()['linewidth'])
		axes[idx].set_ylabel(metric)
		axes[idx].set_yticks(yticks[idx][0])
		axes[idx].set_yticklabels(yticks[idx][1])
		plt.setp(axes[idx].get_xticklabels(), rotation=45, ha='right',
				rotation_mode="anchor")
		axes[idx].set_ylim(yticks[idx][0][0], yticks[idx][0][-1])

	fig.tight_layout()
	this_file_dir = os.path.dirname(os.path.abspath(__file__))
	fig.savefig(os.path.join(this_file_dir, f'save_figs/two_sides_barplot{fig_suffix}'), dpi=300)

def add_bar_plot_dots(ax, x, y, c='#737373', jitter=0.1):
    np.random.seed(42)
    noise = np.random.randn(len(x))
    x += jitter*noise
    ax.scatter(x, y, c=c, s=get_dot_size(), marker='.', zorder=5)
    
def add_box_plot_dots(ax, x, y, jitter=0.1, color='k'):
    np.random.seed(42)
    noise = np.random.randn(len(x))
    x += jitter*noise
    ax.scatter(x, y, c=color, \
        	   s=get_box_plot_setting()['marker size'], \
               marker='o', edgecolors=ax.spines['bottom'].get_edgecolor(),
               zorder=5)
    
# Calculate the statistical significance and determine how many stars to use
def add_significance_symbol_barplot(ax, df, our_method_index, n_group, n_fold):
	methods = list(df.index)
	values = np.asarray(df)
	values = values.reshape(-1, n_group, n_fold)
	width = 1. / len(methods) * 0.8
	for i in range(n_group):
		values_exd = values.copy()
		values_exd[our_method_index, i, :] = -np.inf
		next_best_method_index = np.argmax(values_exd[:, i, :].mean(axis=1))
		our_method_scores = values[our_method_index, i, :]
		next_best_method_scores = values[next_best_method_index, i, :]
		p_values = stats.ttest_rel(our_method_scores, next_best_method_scores, alternative='greater')
		p_values = p_values.pvalue
		stars = get_star_from_pvalue(p_values)
		star_y = our_method_scores.max() * 1.01
		star_x = i + width * (len(methods) - 0.5) * 1. / 2 - 0.1
		ax.text(star_x, star_y, stars, fontsize=LARGE_SIZE, ha='center')
  
def add_significance_symbol_boxplot(ax, df, our_method_index):
	methods = list(df.index)
	values = np.asarray(df)
	for i in range(len(methods)):
		if i == our_method_index:
			continue
		our_method_scores = values[our_method_index, :]
		other_method_scores = values[i, :]
		p_values = stats.ttest_rel(our_method_scores, other_method_scores, alternative='greater')
		p_values = p_values.pvalue
		stars = get_star_from_pvalue(p_values)
		if len(stars) == 0:
			continue
		t = 0.05
		delta_y = i*t
		line_y = our_method_scores.max() + delta_y
		star_x = i + 1
		ax.plot([our_method_index + 1, our_method_index + 1], [line_y - 0.4*t, line_y], c=ax.spines['bottom'].get_edgecolor(), linewidth=1)
		ax.plot([star_x, star_x], [other_method_scores.max() + t, line_y], c=ax.spines['bottom'].get_edgecolor(), linewidth=1)
		ax.plot([our_method_index + 1, star_x], [line_y, line_y], c=ax.spines['bottom'].get_edgecolor(), linewidth=1)
		ax.text((our_method_index + 1 + star_x)/2, line_y - 0.4*t, stars, fontsize=LARGE_SIZE, ha='center')
  

# Calculate the statistical significance and determine how many stars to use
def add_significance_symbol_barplot_mean_std(ax, df, our_method_index, n_group, n_fold=3):
	methods = list(df.index)
	values = np.asarray(df)
	values = values.reshape(-1, n_group, 2)
	width = 1. / len(methods) * 0.8
	for i in range(n_group):
		values_exd = values.copy()
		values_exd[our_method_index, i, 0] = -np.inf
		next_best_method_index = np.argmax(values_exd[:, i, 0])
		our_method_scores = values[our_method_index, i, :]
		next_best_method_scores = values[next_best_method_index, i, :]
		p_values = stats.ttest_ind_from_stats(our_method_scores[0], our_method_scores[1],
                                  n_fold,
                                  next_best_method_scores[0], next_best_method_scores[1],
								  n_fold)
		p_values = p_values.pvalue / 2
		stars = get_star_from_pvalue(p_values)
		star_y = our_method_scores.max() * 1.01
		star_x = i + width * (len(methods) - 0.5) * 1. / 2 - 0.1
		ax.text(star_x, star_y, stars, fontsize=LARGE_SIZE, ha='center')	

def get_star_from_pvalue(p_value):
	if p_value < 0.001:
		stars = '***'
	elif p_value < 0.01:
		stars = '**'
	elif p_value < 0.05:
		stars = '*'
	else:
		stars = ''
	return stars
 
def bar_plot(df, ax, n_group, n_fold, xtick_label, capsize=2, \
	     	capthick=0.5, edgecolor='black', linewidth=0.5):
	"""
	Plot bar plots
	:param df: dataframe, columns: n_group*n_fold, index: methods
	:param axes: axes
	:param n_group: number of groups
	:param n_fold: number of folds
	:param xlabels: x labels
	:param ylabels: y labels
	:return:
	"""
	# calculate the mean and std of each group
	methods = list(df.index)
	values = np.asarray(df)
	values = values.reshape(-1, n_group, n_fold)
	mean_values = np.mean(values, axis=2)
	std_values = np.std(values, axis=2)
	# sort methods, mean_values, std_values according to the mean_values
	'''indices = np.argsort(np.mean(mean_values, axis=1), axis=0)
	methods = [methods[i] for i in indices]
	mean_values = mean_values[indices, :]
	std_values = std_values[indices, :]
	values = values[indices, :, :]'''
	# plot
	width = 1. / len(methods) * 0.8
	for i in range(len(methods)):
		ax.bar(np.arange(n_group) + i * width, mean_values[i, :], \
	 			width, color=get_model_colors(methods[i]), label=methods[i], \
				edgecolor=edgecolor, linewidth=linewidth)
		ax.errorbar(np.arange(n_group) + i * width, mean_values[i, :], \
					yerr=std_values[i, :], color=ax.spines['bottom'].get_edgecolor(), capsize=capsize, \
					capthick=capthick, fmt='none', elinewidth=capthick)
	dots_x, dots_y = [], []
	for i in range(n_group):
		for j in range(n_fold):
			dots_x.extend(np.arange(len(methods))* width + i)
			dots_y.extend(values[:, i, j])
	add_bar_plot_dots(ax, dots_x, dots_y, jitter=width*0.1)
	ax.spines['right'].set_visible(False)
	ax.spines['top'].set_visible(False)
	ax.set_xticklabels(xtick_label)
	ax.set_xticks(np.arange(n_group) + width * (len(methods) - 0.5) * 1. / 2 - 0.1)
 
def bar_plot_exclude_graphsynergy(df, ax, n_group, n_fold, xtick_label, capsize=2, \
	     	capthick=0.5, edgecolor='black', linewidth=0.5):
    """
    Plot bar plots
    :param df: dataframe, columns: n_group*n_fold, index: methods
    :param axes: axes
    :param n_group: number of groups
    :param n_fold: number of folds
    :param xlabels: x labels
    :param ylabels: y labels
    :return:
    """
    # calculate the mean and std of each group
    methods = list(df.index)
    values = np.asarray(df)
    values = values.reshape(-1, n_group, n_fold)
    mean_values = np.mean(values, axis=2)
    std_values = np.std(values, axis=2)
    # sort methods, mean_values, std_values according to the mean_values
    '''indices = np.argsort(np.mean(mean_values, axis=1), axis=0)
    methods = [methods[i] for i in indices]
    mean_values = mean_values[indices, :]
    std_values = std_values[indices, :]
    values = values[indices, :, :]'''
    # plot
    width = 1. / len(methods) * 0.8
    for g in range(n_group):
        if g < 2:
            for i in range(len(methods)):
                if g == 0:
                    ax.bar(g + i * width, mean_values[i, g], \
                            width, color=get_model_colors(methods[i]), label=methods[i], \
                            edgecolor=edgecolor, linewidth=linewidth)
                else:
                    ax.bar(g + i * width, mean_values[i, g], \
                            width, color=get_model_colors(methods[i]), \
                            edgecolor=edgecolor, linewidth=linewidth)
                ax.errorbar(g + i * width, mean_values[i, g], \
                            yerr=std_values[i, g], color=ax.spines['bottom'].get_edgecolor(), capsize=capsize, \
                            capthick=capthick, fmt='none', elinewidth=capthick)
        else:
            new_methods, new_mean, new_err = [], [], []
            for m_i in range(len(methods)):
                if methods[m_i] != 'GraphSynergy':
                    new_methods.append(methods[m_i])
                    new_mean.append(mean_values[m_i, g])
                    new_err.append(std_values[m_i, g])
            for i in range(len(methods)-1):
                ax.bar(g + i * width, new_mean[i], \
                        width, color=get_model_colors(new_methods[i]), \
                        edgecolor=edgecolor, linewidth=linewidth)
                ax.errorbar(g + i * width, new_mean[i], \
                            yerr=new_err[i], color=ax.spines['bottom'].get_edgecolor(), capsize=capsize, \
                            capthick=capthick, fmt='none', elinewidth=capthick)
    dots_x, dots_y = [], []
    for i in range(n_group):
        for j in range(n_fold):
            if i < 2:
                dots_x.extend(np.arange(len(methods))* width + i)
                dots_y.extend(values[:, i, j])
            else:
                dots_x.extend(np.array([0,1,2,3,4])* width + i)
                dots_y.extend([values[0, i, j], values[2, i, j], values[3, i, j], values[4, i, j], values[5, i, j]])
    add_bar_plot_dots(ax, dots_x, dots_y, jitter=width*0.1)
    ax.spines['right'].set_visible(False)
    ax.spines['top'].set_visible(False)
    ax.set_xticklabels(xtick_label)
    ax.set_xticks(np.arange(n_group) + width * (len(methods) - 0.5) * 1. / 2 - 0.1)

def bar_plot_mean_std(df, ax, n_group, xtick_label, capsize=2, \
	     	capthick=0.5, edgecolor='black', linewidth=0.5):
	"""
	Plot bar plots
	:param df: dataframe, columns: n_group*n_fold, index: methods
	:param axes: axes
	:param n_group: number of groups
	:param n_fold: number of folds
	:param xlabels: x labels
	:param ylabels: y labels
	:return:
	"""
	# calculate the mean and std of each group
	methods = list(df.index)
	values = np.asarray(df)
	values = values.reshape(-1, n_group, 2)
	mean_values = values[:, :, 0]
	std_values = values[:, :, 1]
	width = 1. / len(methods) * 0.8
	for i in range(len(methods)):
		ax.bar(np.arange(n_group) + i * width, mean_values[i, :], \
				width, color=get_model_colors(methods[i]), label=methods[i], \
				edgecolor=edgecolor, linewidth=linewidth)
		ax.errorbar(np.arange(n_group) + i * width, mean_values[i, :], \
					yerr=std_values[i, :], color=ax.spines['bottom'].get_edgecolor(), capsize=capsize, \
					capthick=capthick, fmt='none', elinewidth=capthick)
	ax.spines['right'].set_visible(False)
	ax.spines['top'].set_visible(False)
	ax.set_xticklabels(xtick_label)
	ax.set_xticks(np.arange(n_group) + width * (len(methods) - 0.5) * 1. / 2 - 0.1)
 
def bar_plot_mean_std_per_method(df, ax, capsize=2, \
			capthick=0.5, edgecolor='black', linewidth=0.5):
	# calculate the mean and std of each group
	methods = list(df.index)
	values = np.asarray(df)
	mean_values = values[:, 0]
	std_values = values[:, 1]
	width = 1
	for i in range(len(methods)):
		ax.bar(i, mean_values[i], \
				width, color=get_model_colors(methods[i]), label=methods[i], \
				edgecolor=edgecolor, linewidth=linewidth)
		ax.errorbar(i, mean_values[i], \
					yerr=std_values[i], color=ax.spines['bottom'].get_edgecolor(), capsize=capsize, \
					capthick=capthick, fmt='none', elinewidth=capthick)
	ax.spines['right'].set_visible(False)
	ax.spines['top'].set_visible(False)
	#ax.set_xticks([])
	ax.set_xticklabels(methods)
	ax.set_xticks(np.arange(len(methods)))
 
def scatter_plot(ax, x, y, xlabel='', ylabel='', s_scale=1):
    xy = np.vstack([x, y])
    kde = gaussian_kde(xy)
    density = kde(xy)
    sc = ax.scatter(x, y, s=s_scale*get_scatter_plot_setting()['s'], \
        alpha=get_scatter_plot_setting()['alpha'],
        c=density, cmap=get_scatter_plot_setting()['cmap'])
    plt.colorbar(sc, ax=ax)
    format_ax(ax)
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    
def add_confidence_level(x, y, ax):
    sns.regplot(x=x, y=y, ci=95, scatter=False, ax=ax, color='w')
    
def best_response_scatter_plot(preds, labels, fig_suffix='.png'):
    fig = get_scatter_plot_axis()
    # layout the grids for the scatter plot and the histogram plot
    gs = gridspec.GridSpec(2, 2, height_ratios=[1, 6], width_ratios=[6, 1])
    ax_scatter = plt.subplot(gs[1, 0])
    xlabel=r'Observed BestResponse' + '\n' + r'(minimum $\Delta Vol_t$ for $t \geq 10$d)'
    ylabel=r'Predicted BestResponse' + '\n' + r'(minimum $\Delta Vol_t$ for $t \geq 10$d)'
    scatter_plot(ax_scatter, labels, preds, \
        xlabel=r'Observed BestResponse', \
        ylabel=r'Predicted BestResponse')
    slope, intercept, r_value, p_value, std_err = linregress(labels, preds)
    p_value_str = r'$p < {:.1f} \times 10^{{{:.0f}}}$'.format(p_value / 10 ** np.floor(np.log10(p_value)), np.floor(np.log10(p_value)))
    ax_scatter.text(-1.8, 4.8, r'$y = {:.2f}x + {:.2f}$'.format(slope, intercept), fontsize=SMALLER_SIZE)
    ax_scatter.text(-1.8, 4.3, r'$r = {:.2f}$'.format(r_value), fontsize=SMALLER_SIZE)
    ax_scatter.text(-1.8, 3.8, r'{}'.format(p_value_str), fontsize=SMALLER_SIZE)
    ax_scatter.set_xlim([-2, 5.2])
    ax_scatter.set_ylim([-2, 5.2])
    ax_scatter.set_xticks([-2, 0, 2, 4])
    ax_scatter.set_yticks([-2, 0, 2, 4])
    add_confidence_level(labels, preds, ax_scatter)
    # add the histogram plot
    ax_hist_x = plt.subplot(gs[0, 0], sharex=ax_scatter)
    sns.histplot(labels, ax=ax_hist_x, binwidth=get_scatter_plot_setting()['hist_bin_width'],\
        color=get_scatter_plot_setting()['hist_color'], kde=True, kde_kws={'cut': 3})
    ax_hist_x.set_ylabel('')
    ax_hist_y = plt.subplot(gs[1, 1], sharey=ax_scatter)
    sns.histplot(y=preds, ax=ax_hist_y, \
        binwidth=get_scatter_plot_setting()['hist_bin_width'],\
        color=get_scatter_plot_setting()['hist_color'], kde=True)
    ax_hist_y.set_xlabel('')
    for patch in ax_hist_x.patches:
        patch.set_edgecolor('w')
    for patch in ax_hist_y.patches:
        patch.set_edgecolor('w')
    format_ax(ax_hist_x)
    format_ax(ax_hist_y)
    ax_hist_x.set_yticks([])
    ax_hist_y.set_xticks([])
    ax_hist_x.spines['left'].set_visible(False)
    ax_hist_y.spines['bottom'].set_visible(False)
    plt.setp(ax_hist_x.get_xticklabels(), visible=False)
    plt.setp(ax_hist_y.get_yticklabels(), visible=False)
    fig.tight_layout()
    this_file_dir = os.path.dirname(os.path.abspath(__file__))
    fig.savefig(os.path.join(this_file_dir, f'save_figs/xenograft_best_response_scatter_plot{fig_suffix}'), dpi=300)
    
def days_response_scatter_plot(preds, labels, fig_suffix='.png'):
    fig = get_scatter_plot_axis()
    # layout the grids for the scatter plot and the histogram plot
    gs = gridspec.GridSpec(2, 2, height_ratios=[1, 6], width_ratios=[6, 1])
    ax_scatter = plt.subplot(gs[1, 0])
    scatter_plot(ax_scatter, labels, preds, s_scale=1, \
        xlabel=r'Observed tumor change', \
        ylabel=r'Predicted tumor change')
    # set ax_scatter step size
    ax_scatter.set_xticks([0, 2, 4, 6, 8, 10])
    ax_scatter.set_yticks([0, 2, 4, 6])
    slope, intercept, r_value, p_value, std_err = linregress(labels, preds)
    if p_value == 0.0:
        p_value_str = r'$p < 1 \times 10^{-100}$'
    else:
        p_value_str = r'$p < {:.1f} \times 10^{{{:.0f}}}$'.format(p_value / 10 ** np.floor(np.log10(p_value)), np.floor(np.log10(p_value)))
    ax_scatter.text(4.5, 6.2, r'$y = {:.2f}x + {:.2f}$'.format(slope, intercept), fontsize=SMALLER_SIZE)
    ax_scatter.text(4.5, 5.7, r'$r = {:.2f}$'.format(r_value), fontsize=SMALLER_SIZE)
    ax_scatter.text(4.5, 5.2, r'{}'.format(p_value_str), fontsize=SMALLER_SIZE)
    add_confidence_level(labels, preds, ax_scatter)
    # add the histogram plot
    ax_hist_x = plt.subplot(gs[0, 0], sharex=ax_scatter)
    sns.histplot(labels, ax=ax_hist_x, binwidth=get_scatter_plot_setting()['hist_bin_width'],\
        color=get_scatter_plot_setting()['hist_color'], kde=True, kde_kws={'cut': 3})
    ax_hist_x.set_ylabel('')
    ax_hist_y = plt.subplot(gs[1, 1], sharey=ax_scatter)
    sns.histplot(y=preds, ax=ax_hist_y, \
        binwidth=get_scatter_plot_setting()['hist_bin_width'],\
        color=get_scatter_plot_setting()['hist_color'], kde=True)
    ax_hist_y.set_xlabel('')
    for patch in ax_hist_x.patches:
        patch.set_edgecolor('w')
    for patch in ax_hist_y.patches:
        patch.set_edgecolor('w')
    format_ax(ax_hist_x)
    format_ax(ax_hist_y)
    ax_hist_x.set_yticks([])
    ax_hist_y.set_xticks([])
    ax_hist_x.spines['left'].set_visible(False)
    ax_hist_y.spines['bottom'].set_visible(False)
    plt.setp(ax_hist_x.get_xticklabels(), visible=False)
    plt.setp(ax_hist_y.get_yticklabels(), visible=False)
    fig.tight_layout()
    this_file_dir = os.path.dirname(os.path.abspath(__file__))
    fig.savefig(os.path.join(this_file_dir, f'save_figs/xenograft_days_response_scatter_plot{fig_suffix}'), dpi=300)

def days_response_extrapolation_scatter_plot(preds, labels, fig_suffix='.png'):
    fig = get_scatter_plot_axis()
    # layout the grids for the scatter plot and the histogram plot
    gs = gridspec.GridSpec(2, 2, height_ratios=[1, 6], width_ratios=[6, 1])
    ax_scatter = plt.subplot(gs[1, 0])
    scatter_plot(ax_scatter, labels, preds, \
        xlabel=r'Observed tumor change', \
        ylabel=r'Predicted tumor change')
    slope, intercept, r_value, p_value, std_err = linregress(labels, preds)
    p_value_str = r'$p < {:.1f} \times 10^{{{:.0f}}}$'.format(p_value / 10 ** np.floor(np.log10(p_value)), np.floor(np.log10(p_value)))
    ax_scatter.text(-1.6, 4.3, r'$y = {:.2f}x + {:.2f}$'.format(slope, intercept), fontsize=SMALLER_SIZE)
    ax_scatter.text(-1.6, 3.8, r'$r = {:.2f}$'.format(r_value), fontsize=SMALLER_SIZE)
    ax_scatter.text(-1.6, 3.3, r'{}'.format(p_value_str), fontsize=SMALLER_SIZE)
    ax_scatter.set_xlim([-2, 10])
    ax_scatter.set_ylim([-2, 5.0])
    ax_scatter.set_xticks([-1, 1, 3, 5, 7, 9])
    ax_scatter.set_yticks([-1, 1, 3, 5])
    add_confidence_level(labels, preds, ax_scatter)
    # add the histogram plot
    ax_hist_x = plt.subplot(gs[0, 0], sharex=ax_scatter)
    sns.histplot(labels, ax=ax_hist_x, binwidth=get_scatter_plot_setting()['hist_bin_width'],\
        color=get_scatter_plot_setting()['hist_color'], kde=True, kde_kws={'cut': 3})
    ax_hist_x.set_ylabel('')
    ax_hist_y = plt.subplot(gs[1, 1], sharey=ax_scatter)
    sns.histplot(y=preds, ax=ax_hist_y, \
        binwidth=get_scatter_plot_setting()['hist_bin_width'],\
        color=get_scatter_plot_setting()['hist_color'], kde=True)
    ax_hist_y.set_xlabel('')
    for patch in ax_hist_x.patches:
        patch.set_edgecolor('w')
    for patch in ax_hist_y.patches:
        patch.set_edgecolor('w')
    format_ax(ax_hist_x)
    format_ax(ax_hist_y)
    ax_hist_x.set_yticks([])
    ax_hist_y.set_xticks([])
    ax_hist_x.spines['left'].set_visible(False)
    ax_hist_y.spines['bottom'].set_visible(False)
    plt.setp(ax_hist_x.get_xticklabels(), visible=False)
    plt.setp(ax_hist_y.get_yticklabels(), visible=False)
    fig.tight_layout()
    this_file_dir = os.path.dirname(os.path.abspath(__file__))
    fig.savefig(os.path.join(this_file_dir, f'save_figs/xenograft_days_response_extrapolation_scatter_plot{fig_suffix}'), dpi=300)
    
def format_legend(fig_legend, handles, labels, legendmarker=20, loc='center', ncols=1, **kwargs):
    return fig_legend.legend(handles, labels, loc=loc, scatterpoints=1, ncol=ncols,
                      frameon=False, markerscale=legendmarker, **kwargs)

def kaplan_meier_curve(ax, times, observations, labels, xlabel, ylabel, showCI=True,
                       max_time=None, min_y=0.2, usePercentageX=False, usePercentageY=True, colors=None, alpha=0.05):
    kmf1 = KaplanMeierFitter()  # instantiate the class to create an object

    if type(colors) == str:
        colors = [colors for _ in range(len(times))]  # all the same color
    data = []
    for cohort in range(len(times)):
        dt = np.nan_to_num(times[cohort])
        for t, event in zip(dt, observations[cohort]):
            data.append([t, int(event), cohort])
    data = pd.DataFrame(data, columns=['duration', 'event', 'cohort'])
    cph = CoxPHFitter()
    cph.fit(data, duration_col='duration', event_col='event', formula="cohort")
    hazard_ratios = cph.summary["exp(coef)"]
    hr_text = f"\nHazard Ratio: {hazard_ratios[0]:.2f}"
            
    for cohort in range(len(times)):
        dt = np.nan_to_num(times[cohort])
        kmf1.fit(dt, np.nan_to_num(observations[cohort]))
        kmf1.plot(ax=ax, label=labels[cohort], ci_show=showCI, color=colors[cohort] if colors is not None else None,
                  alpha=alpha, linewidth=2)
	
    p = logrank_test(times[0], times[1], observations[0], observations[1]).p_value
	# show p-value in this figure
    ax.text(0.55, 0.8, r'log rank $p$-value: {:.4f}'.format(p), transform=ax.transAxes, \
	    	verticalalignment='top')
    print('p-value: {:.4f}'.format(p))

    if usePercentageX:
        ax.xaxis.set_major_formatter(mtick.PercentFormatter(
            xmax=1, decimals=None, symbol='%', is_latex=False))
    if usePercentageY:
        ax.yaxis.set_major_formatter(mtick.PercentFormatter(
            xmax=1, decimals=None, symbol='%', is_latex=False))

    if max_time is not None:
        ax.set_xlim(ax.get_xlim()[0], max_time)
    ax.set_ylim(min_y, ax.get_ylim()[1])

    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    handles, labels = ax.get_legend_handles_labels()
    labels[0] += hr_text
    leg = format_legend(ax, handles, labels, loc='lower left', handleheight=2.5)
    # align the marker with the first row of text
    text = leg.get_texts()[0]
    text.set_y(text.get_position()[1] - 7)
    
def kaplan_meier_curve_3_cls(ax, times, observations, labels, xlabel, ylabel, showCI=True,
                       max_time=None, min_y=0.2, usePercentageX=False, usePercentageY=True, colors=None, alpha=0.05):
    kmf1 = KaplanMeierFitter()  # instantiate the class to create an object

    if type(colors) == str:
        colors = [colors for _ in range(len(times))]  # all the same color
    data = []
    for cohort in range(2):
        dt = np.nan_to_num(times[cohort])
        for t, event in zip(dt, observations[cohort]):
            data.append([t, int(event), cohort])
    data = pd.DataFrame(data, columns=['duration', 'event', 'cohort'])
    cph = CoxPHFitter()
    cph.fit(data, duration_col='duration', event_col='event', formula="cohort")
    hazard_ratios = cph.summary["exp(coef)"]
    hr_text = f"\nH. R. vs. Pathway not activated (ER+): {hazard_ratios[0]:.2f}"
            
    for cohort in range(len(times)):
        dt = np.nan_to_num(times[cohort])
        kmf1.fit(dt, np.nan_to_num(observations[cohort]))
        kmf1.plot(ax=ax, label=labels[cohort], ci_show=showCI, color=colors[cohort] if colors is not None else None,
                  alpha=alpha, linewidth=2)
    all_times = times[0] + times[1] + times[2]
    all_obs = observations[0] + observations[1] + observations[2]
    all_groups = [0 for _ in range(len(times[0]))] + [1 for _ in range(len(times[1]))] + [2 for _ in range(len(times[2]))]
    p = multivariate_logrank_test(all_times, all_groups, all_obs).p_value
	# show p-value in this figure
    ax.text(0.55, 0.8, r'log rank $p$-value: {:.4f}'.format(p), transform=ax.transAxes, \
	    	verticalalignment='top')
    print('p-value: {:.4f}'.format(p))

    if usePercentageX:
        ax.xaxis.set_major_formatter(mtick.PercentFormatter(
            xmax=1, decimals=None, symbol='%', is_latex=False))
    if usePercentageY:
        ax.yaxis.set_major_formatter(mtick.PercentFormatter(
            xmax=1, decimals=None, symbol='%', is_latex=False))

    if max_time is not None:
        ax.set_xlim(ax.get_xlim()[0], max_time)
    ax.set_ylim(min_y, ax.get_ylim()[1])

    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    handles, labels = ax.get_legend_handles_labels()
    labels[0] += hr_text
    leg = format_legend(ax, handles, labels, loc='lower left', handleheight=2.5, fontsize=10)
    # align the marker with the first row of text
    text = leg.get_texts()[0]
    text.set_y(text.get_position()[1] - 7)

def gdsc_eval_modalities_plot(m2metrics, tgt, tgt_error, metrics='', fig_suffix='.png'):
    xtick_labels = ['SMILES + Graph \n (100000 iterations)']
    for m in m2metrics.keys():
        if m == 'SMILES+Graph':
            continue
        else:
            x_l = '+ {} \n (20000 iterations)'.format(m.split('+')[-1])
            xtick_labels.append(x_l)
    fig, ax = get_broken_line_axis()
    x, y, y_err = [], [], []
    for idx, m in enumerate(m2metrics.keys()):
        x.append(idx)
        y.append(np.mean(m2metrics[m]))
        y_err.append(np.std(m2metrics[m]))
    ax.plot(x, y, color=get_broken_line_setting()['color'], \
        marker=get_broken_line_setting()['marker'], \
        markersize=get_broken_line_setting()['markersize'], \
        markerfacecolor=get_broken_line_setting()['markerfacecolor'], \
        markeredgecolor=get_broken_line_setting()['markeredgecolor'], \
        linewidth=get_broken_line_setting()['linewidth'],
        label='Simultaneous training')
    ax.errorbar(x, y, yerr=y_err, fmt='none', \
        elinewidth=get_broken_line_setting()['error_bar_width'], \
        ecolor=get_broken_line_setting()['error_bar_color'], 
        capsize=get_broken_line_setting()['capsize'])
    ax.plot([x[0], x[-1]], [tgt, tgt], color=get_broken_line_setting()['target_color'], \
        linewidth=get_broken_line_setting()['target_linewidth'],
        linestyle=get_broken_line_setting()['target_linestyle'], label='Incremental training')
    ax.fill_between([x[0], x[-1]], [tgt - tgt_error, tgt - tgt_error], \
        [tgt + tgt_error, tgt + tgt_error], \
        color=get_broken_line_setting()['target_color'], alpha=0.2)
    handles, labels = ax.get_legend_handles_labels()
    fig.legend(handles, labels, loc='upper center', ncol=6, frameon=False)
    ax.set_xticks(x)
    ax.set_xticklabels(xtick_labels, rotation=45, ha='right')
    ax.set_ylabel(metrics)
    format_ax(ax)
    fig.tight_layout()
    this_file_dir = os.path.dirname(os.path.abspath(__file__))
    fig.savefig(os.path.join(this_file_dir, f'save_figs/gdsc_eval_modalities_{metrics}{fig_suffix}'), dpi=300)

def umap_per_drug_combo(names, umap_array, fig_suffix='.png'):
    fig, ax = get_umap_plot_axis()
    cluster_embs = collections.OrderedDict()
    for i in range(len(names)):
        cls_name = '+'.join(names[i].split('+')[1:-2])
        t = eval(names[i].split('+')[-2])
        if cls_name not in cluster_embs:
            cluster_embs[cls_name] = []
        cluster_embs[cls_name].append(umap_array[i, :].reshape(-1, 2))
    for cls_name in cluster_embs:
        cluster_embs[cls_name] = np.concatenate(cluster_embs[cls_name], axis=0)
    for idx, cls_name in enumerate(cluster_embs):
        ax.scatter(cluster_embs[cls_name][:, 0], cluster_embs[cls_name][:, 1], \
            label=cls_name, s=get_umap_plot_setting()['s'], \
            alpha=get_umap_plot_setting()['alpha'], \
            linewidths=get_umap_plot_setting()['marker_linewidth'], \
            c=np.asarray(get_umap_colors(idx)))
    arrow_axes(ax, xlabel='UMAP 1', ylabel='UMAP 2')
    format_ax(ax)
    handles, labels = ax.get_legend_handles_labels()
    ax.legend(handles, labels, loc='upper left', bbox_to_anchor=(1, 1), ncol=1, frameon=False, fontsize=0.6*SMALLER_SIZE)
    ax.set_xticklabels([])
    ax.set_yticklabels([])
    fig.tight_layout()
    this_file_dir = os.path.dirname(os.path.abspath(__file__))
    fig.savefig(os.path.join(this_file_dir, f'save_figs/xenograft_days_response_umap/drug_pair{fig_suffix}'), dpi=300)
    
def umap_per_model(names, umap_array, fig_suffix='.png'):
    fig, ax = get_umap_plot_axis()
    cluster_embs = collections.OrderedDict()
    for i in range(len(names)):
        model_name = names[i].split('+')[0]
        t = eval(names[i].split('+')[-1])
        if model_name not in cluster_embs:
            cluster_embs[model_name] = []
        cluster_embs[model_name].append(umap_array[i, :].reshape(-1, 2))
    sel_names, cluster_n = [], []
    for model_name in cluster_embs:
        sel_names.append(model_name)
        cluster_embs[model_name] = np.concatenate(cluster_embs[model_name], axis=0)
        cluster_n.append(cluster_embs[model_name].shape[0])
    sel_names = np.asarray(sel_names)
    cluster_n = np.asarray(cluster_n)
    srt_idx = np.argsort(cluster_n)[::-1]
    sel_names = sel_names[srt_idx[0: 20]]

    for idx, model_name in enumerate(sel_names):
        ax.scatter(cluster_embs[model_name][:, 0], cluster_embs[model_name][:, 1], \
            label=model_name, s=get_umap_plot_setting()['s'], \
            alpha=get_umap_plot_setting()['alpha'], \
            linewidths=get_umap_plot_setting()['marker_linewidth'], \
            c=np.asarray(get_umap_colors(idx)))
    ax.set_xlabel('UMAP 1')
    ax.set_ylabel('UMAP 2')
    format_ax(ax)
    handles, labels = ax.get_legend_handles_labels()
    ax.legend(handles, labels, loc='upper left', bbox_to_anchor=(1, 1), ncol=1, frameon=False, fontsize=0.6*SMALLER_SIZE)
    ax.set_xticklabels([])
    ax.set_yticklabels([])
    fig.tight_layout()
    this_file_dir = os.path.dirname(os.path.abspath(__file__))
    fig.savefig(os.path.join(this_file_dir, f'save_figs/xenograft_days_response_umap_per_model{fig_suffix}'), dpi=300)
    
def umap_per_fold(names, umap_array, folds, fig_suffix='.png'):
    fig, ax = get_umap_plot_axis()
    cluster_embs = collections.OrderedDict()
    for i in range(len(folds)):
        fold_i= folds[i]
        if fold_i not in cluster_embs:
            cluster_embs[fold_i] = []
        cluster_embs[fold_i].append(umap_array[i, :].reshape(-1, 2))
    for fold_i in cluster_embs:
        cluster_embs[fold_i] = np.concatenate(cluster_embs[fold_i], axis=0)

    for idx, fold_i in enumerate(cluster_embs):
        ax.scatter(cluster_embs[fold_i][:, 0], cluster_embs[fold_i][:, 1], \
            label=fold_i, s=get_umap_plot_setting()['s'], \
            alpha=get_umap_plot_setting()['alpha'], \
            linewidths=get_umap_plot_setting()['marker_linewidth'], \
            c=np.asarray(get_umap_colors(idx)))
    ax.set_xlabel('UMAP 1')
    ax.set_ylabel('UMAP 2')
    format_ax(ax)
    handles, labels = ax.get_legend_handles_labels()
    ax.legend(handles, labels, loc='upper left', bbox_to_anchor=(1, 1), ncol=1, frameon=False, fontsize=0.6*SMALLER_SIZE)
    ax.set_xticklabels([])
    ax.set_yticklabels([])
    fig.tight_layout()
    this_file_dir = os.path.dirname(os.path.abspath(__file__))
    fig.savefig(os.path.join(this_file_dir, f'save_figs/xenograft_days_response_umap_per_fold{fig_suffix}'), dpi=300)

def umap_per_dvolt(names, umap_array, labels, fig_suffix='.png'):
    fig, ax = get_umap_plot_axis()
    
    sc = ax.scatter(umap_array[:, 0], umap_array[:, 1], \
		s=get_umap_plot_setting()['s'], \
		alpha=get_umap_plot_setting()['alpha'], \
		linewidths=get_umap_plot_setting()['marker_linewidth'], \
		c=labels, cmap=get_umap_plot_setting()['cmap'])
    format_ax(ax)
    plt.colorbar(sc, ax=ax, label=r'$\Delta Vol_{t}$')
    ax.set_xticklabels([])
    ax.set_yticklabels([])
    ax.set_xticks([])
    ax.set_yticks([])
    arrow_axes(ax, xlabel='UMAP 1', ylabel='UMAP 2')
    fig.tight_layout()
    this_file_dir = os.path.dirname(os.path.abspath(__file__))
    fig.savefig(os.path.join(this_file_dir, f'save_figs/xenograft_days_response_umap/dvolt{fig_suffix}'), dpi=300)

def umap_per_time(names, umap_array, labels, fig_suffix='.png'):
    fig, ax = get_umap_plot_axis()
    times = []
    for name in names:
        times.append(eval(name.split('+')[-1]))
    sc = ax.scatter(umap_array[:, 0], umap_array[:, 1], \
		s=get_umap_plot_setting()['s'], \
		alpha=get_umap_plot_setting()['alpha'], \
		linewidths=get_umap_plot_setting()['marker_linewidth'], \
		c=times, cmap=get_umap_plot_setting()['cmap'])
    ax.set_xlabel('UMAP 1')
    ax.set_ylabel('UMAP 2')
    format_ax(ax)
    plt.colorbar(sc, ax=ax, label='Time (days)')
    ax.set_xticklabels([])
    ax.set_yticklabels([])
    fig.tight_layout()
    this_file_dir = os.path.dirname(os.path.abspath(__file__))
    fig.savefig(os.path.join(this_file_dir, f'save_figs/xenograft_days_response_umap_time{fig_suffix}'), dpi=300)

def load_simulations(fold=0):
    path='/home/swang/xuhw/research-projects/Pisces/Pisces/scripts/case_study/xenograft/output/simulate_umap_emb.pkl'
    simulate_embs = load_obj(path)
    return simulate_embs[f"fold{fold}"]

def quantile_normalize(z):
    if len(z) == 0:
        return np.array([])
    srt_idx = np.argsort(z)
    results = np.zeros(len(z))
    for i in range(len(z)):
        results[srt_idx[i]] = i / len(z)
    return results

def velocity_plot_per_time(names, umap_array, preds, labels, 
                            n_time=20, n_clusters=1,
                            eps=1.3, min_samples=5,
                            clustering_method='dbscan',
                            fig_suffix='.png'):
    
    fig, ax = get_umap_plot_axis()
    cmb_time_umaps = calculate_velocity(names, umap_array, preds, labels)
    sel_cmbs = []
    for cmb in cmb_time_umaps:
        if len(cmb_time_umaps[cmb]['time']) > n_time:
            sel_cmbs.append(cmb)
    for cmb in sel_cmbs:
        adata = anndata.AnnData(cmb_time_umaps[cmb]['umap'])
        adata.obsm['X_umap'] = cmb_time_umaps[cmb]['umap']
        adata.obs['dvolt'] = cmb_time_umaps[cmb]['label']
        velocity_plot_utils.velocity_embedding_stream(
            adata, V=cmb_time_umaps[cmb]['vel'], 
            n_neighbors=10, color='dvolt', cutoff_perc=0.0,
            alpha=0.3, density=0.05, smooth=0.9,
            size=200, linewidth=2, arrowsize=2, ax=ax,
            arrow_color='whitesmoke', 
            color_map='magma_r', legend_loc='upper right')
    
    #sim_emb = load_simulations(fold)
    if clustering_method == 'kmeans':
        clusters = KMeans(n_clusters=n_clusters, random_state=0).fit(umap_array)
    elif clustering_method == 'dbscan':
        clusters = DBSCAN(eps=eps, min_samples=min_samples).fit(umap_array)
        n_clusters = np.max(clusters.labels_)+1
    else:
        raise ValueError(f"Clustering method {clustering_method} not supported!")
    n_grid=100
    for i in range(n_clusters):
        idx = clusters.labels_ == i
        x = umap_array[idx, 0]
        y = umap_array[idx, 1]
        z = preds[idx].astype(np.float)
        z = quantile_normalize(z)
        grid_x, grid_y = np.meshgrid(np.linspace(min(x), max(x), n_grid), np.linspace(min(y), max(y), n_grid))
        grid_height = interpolate.griddata((x, y), z, (grid_x, grid_y), method='linear')
        grid_height = gaussian_filter(grid_height, sigma=1.2)
        ax.contour(grid_x, grid_y, grid_height, colors='lightgrey', linewidths=1, linestyles='dashed', levels=7)
    
    sc = ax.scatter(umap_array[:, 0], umap_array[:, 1], \
		s=get_umap_plot_setting()['s'], \
		alpha=get_umap_plot_setting()['alpha'], \
		linewidths=get_umap_plot_setting()['marker_linewidth'], \
		c=labels, cmap=get_umap_plot_setting()['cmap'])
    format_ax(ax)
    plt.colorbar(sc, ax=ax, label=r'$\Delta Vol_{t}$')
    ax.set_xticklabels([])
    ax.set_yticklabels([])
    ax.set_xticks([])
    ax.set_yticks([])
    arrow_axes(ax, xlabel='UMAP 1', ylabel='UMAP 2')
    fig.tight_layout()

    this_file_dir = os.path.dirname(os.path.abspath(__file__))
    fig.savefig(os.path.join(this_file_dir, f'save_figs/xenograft_days_response_velocity/dvolt{fig_suffix}'), dpi=300)

def velocity_plot_per_dvolt(names, umap_array, preds, labels, 
                            n_time=20, n_clusters=1,
                            eps=1.3, min_samples=5,
                            contour_obj='time',
                            clustering_method='dbscan',
                            fig_suffix='.png'):
    
    fig, ax = get_umap_plot_axis()
    cmb_time_umaps = calculate_velocity(names, umap_array, preds, labels)
    sel_cmbs = []
    for cmb in cmb_time_umaps:
        if len(cmb_time_umaps[cmb]['time']) > n_time:
            sel_cmbs.append(cmb)
    for cmb in sel_cmbs:
        adata = anndata.AnnData(cmb_time_umaps[cmb]['umap'])
        adata.obsm['X_umap'] = cmb_time_umaps[cmb]['umap']
        adata.obs['dvolt'] = cmb_time_umaps[cmb]['label']
        velocity_plot_utils.velocity_embedding_stream(
            adata, V=cmb_time_umaps[cmb]['vel'], 
            n_neighbors=10, color='dvolt', cutoff_perc=0.0,
            alpha=0.3, density=0.05, smooth=0.9,
            size=200, linewidth=2, arrowsize=1, ax=ax,
            arrow_color='whitesmoke', 
            color_map='magma_r', legend_loc='upper right')
    
    if contour_obj == 'time':
        times = []
        for i in range(len(names)):
            t = eval(names[i].split('+')[3])
            times.append(t)
        times = np.array(times)

    if clustering_method == 'kmeans':
        clusters = KMeans(n_clusters=n_clusters, random_state=0).fit(umap_array)
    elif clustering_method == 'dbscan':
        clusters = DBSCAN(eps=eps, min_samples=min_samples).fit(umap_array)
        n_clusters = np.max(clusters.labels_)+1
    else:
        raise ValueError(f"Clustering method {clustering_method} not supported!")
    n_grid=100
    for i in range(n_clusters):
        idx = clusters.labels_ == i
        x = umap_array[idx, 0]
        y = umap_array[idx, 1]
        if contour_obj == 'time':
            z = times[idx]
        elif contour_obj == 'dvolt':
            z = preds[idx].astype(np.float)
        z = quantile_normalize(z)
        grid_x, grid_y = np.meshgrid(np.linspace(min(x), max(x), n_grid), np.linspace(min(y), max(y), n_grid))
        grid_height = interpolate.griddata((x, y), z, (grid_x, grid_y), method='linear')
        grid_height = gaussian_filter(grid_height, sigma=1.2)
        ax.contour(grid_x, grid_y, grid_height, colors='#737373', linewidths=1, linestyles='dashed', levels=7)
    
    sc = ax.scatter(umap_array[:, 0], umap_array[:, 1], \
		s=get_umap_plot_setting()['s'], \
		alpha=get_umap_plot_setting()['alpha'], \
		linewidths=get_umap_plot_setting()['marker_linewidth'], \
		c=labels, cmap=get_umap_plot_setting()['cmap'])
    format_ax(ax)
    plt.colorbar(sc, ax=ax, label=r'$\Delta Vol_{t}$')
    ax.set_xticklabels([])
    ax.set_yticklabels([])
    ax.set_xticks([])
    ax.set_yticks([])
    arrow_axes(ax, xlabel='UMAP 1', ylabel='UMAP 2')
    fig.tight_layout()

    this_file_dir = os.path.dirname(os.path.abspath(__file__))
    fig.savefig(os.path.join(this_file_dir, f'save_figs/xenograft_days_response_velocity/dvolt{fig_suffix}'), dpi=300)
    
def velocity_plot_per_dvolt_other_fold(ax, names, umap_array, preds, labels, 
                            n_time=20, n_clusters=1,
                            eps=1.3, min_samples=5,
                            contour_obj='time',
                            title='',
                            clustering_method='dbscan'):
    cmb_time_umaps = calculate_velocity(names, umap_array, preds, labels)
    sel_cmbs = []
    for cmb in cmb_time_umaps:
        if len(cmb_time_umaps[cmb]['time']) > n_time:
            sel_cmbs.append(cmb)
    for cmb in sel_cmbs:
        adata = anndata.AnnData(cmb_time_umaps[cmb]['umap'])
        adata.obsm['X_umap'] = cmb_time_umaps[cmb]['umap']
        adata.obs['dvolt'] = cmb_time_umaps[cmb]['label']
        velocity_plot_utils.velocity_embedding_stream(
            adata, V=cmb_time_umaps[cmb]['vel'], 
            n_neighbors=10, color='dvolt', cutoff_perc=0.0,
            alpha=0.3, density=0.05, smooth=0.9,
            size=200, linewidth=2, arrowsize=1, ax=ax,
            arrow_color='whitesmoke', 
            color_map='magma_r', legend_loc='upper right')
    
    if contour_obj == 'time':
        times = []
        for i in range(len(names)):
            t = eval(names[i].split('+')[3])
            times.append(t)
        times = np.array(times)
    
    #sim_emb = load_simulations(fold)
    if clustering_method == 'kmeans':
        clusters = KMeans(n_clusters=n_clusters, random_state=0).fit(umap_array)
    elif clustering_method == 'dbscan':
        clusters = DBSCAN(eps=eps, min_samples=min_samples).fit(umap_array)
        n_clusters = np.max(clusters.labels_)+1
    else:
        raise ValueError(f"Clustering method {clustering_method} not supported!")
    n_grid=100
    for i in range(n_clusters):
        idx = clusters.labels_ == i
        x = umap_array[idx, 0]
        y = umap_array[idx, 1]
        if contour_obj == 'time':
            z = times[idx]
        elif contour_obj == 'dvolt':
            z = preds[idx].astype(np.float)
        z = quantile_normalize(z)
        grid_x, grid_y = np.meshgrid(np.linspace(min(x), max(x), n_grid), np.linspace(min(y), max(y), n_grid))
        grid_height = interpolate.griddata((x, y), z, (grid_x, grid_y), method='linear')
        grid_height = gaussian_filter(grid_height, sigma=1.2)
        ax.contour(grid_x, grid_y, grid_height, colors='#737373', linewidths=1, linestyles='dashed', levels=7)
    
    sc = ax.scatter(umap_array[:, 0], umap_array[:, 1], \
		s=get_umap_plot_setting()['s'], \
		alpha=get_umap_plot_setting()['alpha'], \
		linewidths=get_umap_plot_setting()['marker_linewidth'], \
		c=labels, cmap=get_umap_plot_setting()['cmap'])
    format_ax(ax)
    plt.colorbar(sc, ax=ax, label=r'$\Delta Vol_{t}$')
    ax.set_xticklabels([])
    ax.set_yticklabels([])
    ax.set_xticks([])
    ax.set_yticks([])
    ax.set_title(title)
    arrow_axes(ax, xlabel='UMAP 1', ylabel='UMAP 2')
    y0,y1 = ax.get_ylim()
    x0,x1 = ax.get_xlim()
    ax.set_aspect(abs(x1-x0)/abs(y1-y0))
    
def velocity_plot_per_drug_combo(names, umap_array, preds, labels, 
                                contour_obj='time',
                                n_time=20, n_clusters=1, fig_suffix='.png'):
    fig, ax = get_drug_combo_umap_plot_axis()
    
    cmb_time_umaps = calculate_velocity(names, umap_array, preds, labels)
    sel_cmbs = []
    for cmb in cmb_time_umaps:
        if len(cmb_time_umaps[cmb]['time']) > n_time:
            sel_cmbs.append(cmb)
    '''for cmb in sel_cmbs:
        adata = anndata.AnnData(cmb_time_umaps[cmb]['umap'])
        adata.obsm['X_umap'] = cmb_time_umaps[cmb]['umap']
        adata.obs['dvolt'] = cmb_time_umaps[cmb]['label']
        velocity_plot_utils.velocity_embedding_stream(
            adata, V=cmb_time_umaps[cmb]['vel'], 
            n_neighbors=10, color='dvolt', cutoff_perc=0.0,
            alpha=0.3, density=0.05, smooth=0.9,
            size=200, linewidth=2, arrowsize=1, ax=ax,
            arrow_color='whitesmoke', 
            color_map='magma_r', legend_loc='upper right')'''
    
    if contour_obj == 'time':
        times = []
        for i in range(len(names)):
            t = eval(names[i].split('+')[3])
            times.append(t)
        times = np.array(times)
    
    kmeans = KMeans(n_clusters=n_clusters, random_state=0).fit(umap_array)
    n_grid=100
    for i in range(n_clusters):
        idx = kmeans.labels_ == i
        x = umap_array[idx, 0]
        y = umap_array[idx, 1]
        if contour_obj == 'time':
            z = times[idx]
        elif contour_obj == 'dvolt':
            z = preds[idx].astype(np.float)
        z = quantile_normalize(z)
        grid_x, grid_y = np.meshgrid(np.linspace(min(x), max(x), n_grid), np.linspace(min(y), max(y), n_grid))
        grid_height = interpolate.griddata((x, y), z, (grid_x, grid_y), method='linear')
        grid_height = gaussian_filter(grid_height, sigma=1.2)
        ax.contour(grid_x, grid_y, grid_height, colors='#737373', linewidths=1, linestyles='dashed', levels=7)
    
    cluster_embs = collections.OrderedDict()
    for i in range(len(names)):
        cls_name = '+'.join(names[i].split('+')[1:-2])
        t = eval(names[i].split('+')[-2])
        if cls_name not in cluster_embs:
            cluster_embs[cls_name] = []
        cluster_embs[cls_name].append(umap_array[i, :].reshape(-1, 2))
    for cls_name in cluster_embs:
        cluster_embs[cls_name] = np.concatenate(cluster_embs[cls_name], axis=0)
    for idx, cls_name in enumerate(cluster_embs):
        ax.scatter(cluster_embs[cls_name][:, 0], cluster_embs[cls_name][:, 1], \
            label=cls_name, s=get_umap_plot_setting()['s'], \
            alpha=get_umap_plot_setting()['alpha'], \
            linewidths=get_umap_plot_setting()['marker_linewidth'], \
            c=get_velocity_color_markers(idx)[0], \
            marker=get_velocity_color_markers(idx)[1])
    arrow_axes(ax, xlabel='UMAP 1', ylabel='UMAP 2')
    format_ax(ax)
    handles, labels = ax.get_legend_handles_labels()
    ax.legend(handles, labels, loc='upper left', bbox_to_anchor=(1, 1), ncol=1, frameon=False, fontsize=0.6*SMALLER_SIZE)
    ax.set_xticklabels([])
    ax.set_yticklabels([])
    fig.tight_layout()
    this_file_dir = os.path.dirname(os.path.abspath(__file__))
    fig.savefig(os.path.join(this_file_dir, f'save_figs/xenograft_days_response_velocity/drug_pair{fig_suffix}'), dpi=300)

def velocity_plot_per_drug_combo_other_fold(ax, names, umap_array, preds, labels, 
                                contour_obj='time',
                                n_time=20, n_clusters=1, title=''):
    cmb_time_umaps = calculate_velocity(names, umap_array, preds, labels)
    sel_cmbs = []
    for cmb in cmb_time_umaps:
        if len(cmb_time_umaps[cmb]['time']) > n_time:
            sel_cmbs.append(cmb)
    '''for cmb in sel_cmbs:
        adata = anndata.AnnData(cmb_time_umaps[cmb]['umap'])
        adata.obsm['X_umap'] = cmb_time_umaps[cmb]['umap']
        adata.obs['dvolt'] = cmb_time_umaps[cmb]['label']
        velocity_plot_utils.velocity_embedding_stream(
            adata, V=cmb_time_umaps[cmb]['vel'], 
            n_neighbors=10, color='dvolt', cutoff_perc=0.0,
            alpha=0.3, density=0.05, smooth=0.9,
            size=200, linewidth=2, arrowsize=1, ax=ax,
            arrow_color='whitesmoke', 
            color_map='magma_r', legend_loc='upper right')'''
    
    if contour_obj == 'time':
        times = []
        for i in range(len(names)):
            t = eval(names[i].split('+')[3])
            times.append(t)
        times = np.array(times)
    
    kmeans = KMeans(n_clusters=n_clusters, random_state=0).fit(umap_array)
    n_grid=100
    for i in range(n_clusters):
        idx = kmeans.labels_ == i
        x = umap_array[idx, 0]
        y = umap_array[idx, 1]
        if contour_obj == 'time':
            z = times[idx]
        elif contour_obj == 'dvolt':
            z = preds[idx].astype(np.float)
        z = quantile_normalize(z)
        grid_x, grid_y = np.meshgrid(np.linspace(min(x), max(x), n_grid), np.linspace(min(y), max(y), n_grid))
        grid_height = interpolate.griddata((x, y), z, (grid_x, grid_y), method='linear')
        grid_height = gaussian_filter(grid_height, sigma=1.2)
        ax.contour(grid_x, grid_y, grid_height, colors='#737373', linewidths=1, linestyles='dashed', levels=7)
    
    cluster_embs = collections.OrderedDict()
    for i in range(len(names)):
        cls_name = '+'.join(names[i].split('+')[1:-2])
        t = eval(names[i].split('+')[-2])
        if cls_name not in cluster_embs:
            cluster_embs[cls_name] = []
        cluster_embs[cls_name].append(umap_array[i, :].reshape(-1, 2))
    for cls_name in cluster_embs:
        cluster_embs[cls_name] = np.concatenate(cluster_embs[cls_name], axis=0)
    for idx, cls_name in enumerate(cluster_embs):
        ax.scatter(cluster_embs[cls_name][:, 0], cluster_embs[cls_name][:, 1], \
            label=cls_name, s=get_umap_plot_setting()['s'], \
            alpha=get_umap_plot_setting()['alpha'], \
            linewidths=get_umap_plot_setting()['marker_linewidth'], \
            c=get_velocity_color_markers(idx)[0], \
            marker=get_velocity_color_markers(idx)[1])
    arrow_axes(ax, xlabel='UMAP 1', ylabel='UMAP 2')
    format_ax(ax)
    handles, labels = ax.get_legend_handles_labels()
    ax.legend(handles, labels, loc='upper left', bbox_to_anchor=(1, 1), ncol=1, frameon=False, fontsize=0.6*SMALLER_SIZE)
    ax.set_xticklabels([])
    ax.set_yticklabels([])
    ax.set_title(title)
    y0,y1 = ax.get_ylim()
    x0,x1 = ax.get_xlim()
    ax.set_aspect(abs(x1-x0)/abs(y1-y0))

def velocity_subplot_per_model(names, umap_array, preds, labels, drug1, drug2, n_clusters=1, n_time=20, fig_suffix='.png'):
    fig, ax = get_umap_plot_axis()
    
    cmb_time_umaps = calculate_velocity(names, umap_array, preds, labels)
    sel_cmbs = []
    for cmb in cmb_time_umaps:
        if drug1 in cmb and drug2 in cmb and len(cmb_time_umaps[cmb]['time']) > n_time:
            sel_cmbs.append(cmb)

    kmeans = KMeans(n_clusters=n_clusters, random_state=0).fit(umap_array)
    n_grid=300
    x, y, z = [], [], []
    for cmb in sel_cmbs:
        x.append(cmb_time_umaps[cmb]['umap'][:, 0])
        y.append(cmb_time_umaps[cmb]['umap'][:, 1])
        z.append(cmb_time_umaps[cmb]['time'])
    x = np.concatenate(x, axis=0)
    y = np.concatenate(y, axis=0)
    z = np.concatenate(z, axis=0)
    z = quantile_normalize(z)
    grid_x, grid_y = np.meshgrid(np.linspace(min(x), max(x), n_grid), np.linspace(min(y), max(y), n_grid))
    grid_height = interpolate.griddata((x, y), z, (grid_x, grid_y), method='linear')
    grid_height = gaussian_filter(grid_height, sigma=1)
    ax.contour(grid_x, grid_y, grid_height, colors='#737373', linewidths=1, linestyles='dashed', levels=10)
        
    model_embs = collections.OrderedDict()
    model_labels = collections.OrderedDict()
    for cmb in sel_cmbs:
        model_name = cmb.split('+')[0]
        if model_name not in model_embs:
            model_embs[model_name] = []
            model_labels[model_name] = []
        cmb_umaps = cmb_time_umaps[cmb]['umap']
        model_embs[model_name].append(cmb_umaps)
        model_labels[model_name].append(cmb_time_umaps[cmb]['label'])
    sel_names, cluster_n = [], []
    for model_name in model_embs:
        sel_names.append(model_name)
        model_embs[model_name] = np.concatenate(model_embs[model_name], axis=0)
        model_labels[model_name] = np.concatenate(model_labels[model_name], axis=0)
        cluster_n.append(model_embs[model_name].shape[0])
    sel_names = np.asarray(sel_names)
    cluster_n = np.asarray(cluster_n)

    import matplotlib.colors as colors
    vmin, vmax = np.min(labels), np.max(labels)
    norm = colors.Normalize(vmin=vmin, vmax=vmax)
    for idx, model_name in enumerate(sel_names):
        ax.scatter(model_embs[model_name][:, 0], model_embs[model_name][:, 1], \
            label=model_name, s=3*get_umap_plot_setting()['s'], \
            marker=get_velocity_markers(idx), \
            alpha=1, \
            linewidths=get_umap_plot_setting()['marker_linewidth'], \
            edgecolors=get_umap_plot_setting()['edgecolors'], \
            c=model_labels[model_name], cmap=get_umap_plot_setting()['cmap'], norm=norm)
    x0, x1 = ax.get_xlim()
    y0, y1 = ax.get_ylim()
    ax.set_aspect(abs(x1-x0)/abs(y1-y0))
    format_ax(ax)
    ax.set_xticks([])
    ax.set_yticks([])
    # Remove spines
    ax.spines[['top', 'right', 'bottom', 'left']].set_visible(False)
    handles, labels = ax.get_legend_handles_labels()
    ax.legend(handles, labels, loc='upper left', bbox_to_anchor=(1, 1), ncol=1, frameon=False, fontsize=0.6*SMALLER_SIZE)
    ax.set_xticklabels([])
    ax.set_yticklabels([])
    ax.set_xticks([])
    ax.set_yticks([])
    ax.set_title(f'{drug1} + {drug2}')
    fig.tight_layout()
    this_file_dir = os.path.dirname(os.path.abspath(__file__))
    fig.savefig(os.path.join(this_file_dir, f'save_figs/xenograft_days_response_velocity/model{fig_suffix}'), dpi=300)

def heatmap(data, row_labels, col_labels, ax=None,
			cbar_kw={}, cbarlabel="", **kwargs):
	"""
	Create a heatmap from a numpy array and two lists of labels.
	Parameters
	----------
	data
		A 2D numpy array of shape (N, M).
	row_labels
		A list or array of length N with the labels for the rows.
	col_labels
		A list or array of length M with the labels for the columns.
	ax
		A `matplotlib.axes.Axes` instance to which the heatmap is plotted.  If
		not provided, use current axes or create a new one.  Optional.
	cbar_kw
		A dictionary with arguments to `matplotlib.Figure.colorbar`.  Optional.
	cbarlabel
		The label for the colorbar.  Optional.
	**kwargs
		All other arguments are forwarded to `imshow`.
	"""

	if not ax:
		ax = plt.gca()

	# Plot the heatmap
	im = ax.imshow(data, **kwargs)

	# Create colorbar
	#cbar = ax.figure.colorbar(im, ax=ax, **cbar_kw)
	#cbar.ax.set_ylabel(cbarlabel, rotation=-90, va="bottom")

	# We want to show all ticks...
	ax.set_xticks(np.arange(data.shape[1]))
	ax.set_yticks(np.arange(data.shape[0]))
	# ... and label them with the respective list entries.
	ax.set_xticklabels(col_labels)
	ax.set_yticklabels(row_labels,)

	# Let the horizontal axes labeling appear on top.
	ax.tick_params(top=False, bottom=True,
				   labeltop=False, labelbottom=True)

	# Rotate the tick labels and set their alignment.
	plt.setp(ax.get_xticklabels(), rotation=45,ha="right", va="center",
			 rotation_mode="anchor")#, ha="right"

	# Turn spines off and create white grid.
	#for edge, spine in ax.spines.items():
	#	spine.set_visible(False)

	ax.set_xticks(np.arange(data.shape[1]+1)-.5, minor=True)
	ax.set_yticks(np.arange(data.shape[0]+1)-.5, minor=True)
	ax.grid(which="minor", color="w", linestyle='-', linewidth=2)
	ax.tick_params(which="minor", bottom=False, left=False)

	return im


def annotate_heatmap(im, data=None, valfmt="{x:.2f}",
					 textcolors=["black", "white"],
					 threshold=None, **textkw):
	"""
	A function to annotate a heatmap.
	Parameters
	----------
	im
		The AxesImage to be labeled.
	data
		Data used to annotate.  If None, the image's data is used.  Optional.
	valfmt
		The format of the annotations inside the heatmap.  This should either
		use the string format method, e.g. "$ {x:.2f}", or be a
		`matplotlib.ticker.Formatter`.  Optional.
	textcolors
		A list or array of two color specifications.  The first is used for
		values below a threshold, the second for those above.  Optional.
	threshold
		Value in data units according to which the colors from textcolors are
		applied.  If None (the default) uses the middle of the colormap as
		separation.  Optional.
	**kwargs
		All other arguments are forwarded to each call to `text` used to create
		the text labels.
	"""

	if not isinstance(data, (list, np.ndarray)):
		data = im.get_array()

	# Normalize the threshold to the images color range.
	if threshold is not None:
		threshold = im.norm(threshold)
	else:
		threshold = im.norm(data.max())/2.

	# Set default alignment to center, but allow it to be
	# overwritten by textkw.
	kw = dict(horizontalalignment="center",
			  verticalalignment="center")
	kw.update(textkw)

	# Get the formatter in case a string is supplied
	if isinstance(valfmt, str):
		valfmt = mpl.ticker.StrMethodFormatter(valfmt)

	# Loop over the data and create a `Text` for each "pixel".
	# Change the text's color depending on the data.
	texts = []
	for i in range(data.shape[0]):
		for j in range(data.shape[1]):

			kw.update(color='black')
			text = im.axes.text(j, i, valfmt(data[i, j], None), **kw, fontsize=4)
			texts.append(text)

	return texts

def calculate_velocity(names, umaps, preds, labels):
    cmb_time_umaps = collections.OrderedDict()
    for i in range(len(names)):
        m, d1, d2, t, _ = names[i].split('+')
        cmb_name = f'{m}+' + '+'.join(sorted([d1, d2]))
        if cmb_name not in cmb_time_umaps:
            cmb_time_umaps[cmb_name] = collections.OrderedDict()
            cmb_time_umaps[cmb_name]['time'] = []
            cmb_time_umaps[cmb_name]['umap'] = []
            cmb_time_umaps[cmb_name]['pred'] = []
            cmb_time_umaps[cmb_name]['label'] = []
        cmb_time_umaps[cmb_name]['time'].append(float(t))
        cmb_time_umaps[cmb_name]['umap'].append(umaps[i])
        cmb_time_umaps[cmb_name]['pred'].append(preds[i])
        cmb_time_umaps[cmb_name]['label'].append(labels[i])
    for cmb_name in cmb_time_umaps:
        cmb_time_umaps[cmb_name]['time'] = np.asarray(cmb_time_umaps[cmb_name]['time'])
        srt_idx = np.argsort(cmb_time_umaps[cmb_name]['time'])
        cmb_time_umaps[cmb_name]['time'] = cmb_time_umaps[cmb_name]['time'][srt_idx]
        cmb_time_umaps[cmb_name]['umap'] = np.asarray(cmb_time_umaps[cmb_name]['umap'])[srt_idx, :]
        cmb_time_umaps[cmb_name]['pred'] = np.asarray(cmb_time_umaps[cmb_name]['pred'])[srt_idx]
        cmb_time_umaps[cmb_name]['label'] = np.asarray(cmb_time_umaps[cmb_name]['label'])[srt_idx]
        
        vels = []
        for i in range(len(cmb_time_umaps[cmb_name]['time'])):
            if i == 0 or i == len(cmb_time_umaps[cmb_name]['time']) - 1:
                vel = np.zeros([2])
            else:
                vel = (cmb_time_umaps[cmb_name]['umap'][i+1] - cmb_time_umaps[cmb_name]['umap'][i-1]) / \
                    (cmb_time_umaps[cmb_name]['time'][i+1] - cmb_time_umaps[cmb_name]['time'][i-1])
            vels.append(vel)
        cmb_time_umaps[cmb_name]['vel'] = np.stack(vels)
    return cmb_time_umaps

def eval_topk_plot(df_list, fig_suffix='.png'):
    jitter = 0.1
    fig, axes = get_four_rows_axis()
    for i, k in enumerate(df_list.keys()):
        ax = axes[i]
        df = df_list[k]
        df_array = np.asarray(df)
        dots_x, dots_y = [], []
        for x in range(df_array.shape[0]):
            for y in range(df_array.shape[1]):
                dots_x.append(x)
                dots_y.append(df_array[x, y])
        indices = np.array([0, 1, 2, 3, 4, 5])
        df_mean = np.mean(df_array, axis=1)
        df_err = np.std(df_array, axis=1)
        ax.plot(indices, df_mean, marker='o', color='#5ab4ac')
        ax.errorbar(indices, df_mean, yerr=df_err, fmt='o', color='#5ab4ac', \
                    capsize=8, capthick=1)
        ax.fill_between(indices, df_mean - df_err, df_mean + df_err, \
                        color='#01665e', alpha=0.2)
        format_ax(ax)
        ax.set_xticks(indices)
        ax.set_xticklabels([r'Top 2', r'Top 4', r'Top 8', r'Top 16', r'Top 32', r'Top 64'])
        ax.set_ylabel(k)
        np.random.seed(42)
        noise = np.random.randn(len(dots_x))
        dots_x += jitter*noise
        ax.scatter(dots_x, dots_y, c='w', s=15, marker='.', zorder=5)
        if k != 'BACC':
            ax.set_ylim([0.4, 0.5])
    fig.tight_layout()
    this_file_dir = os.path.dirname(os.path.abspath(__file__))
    fig.savefig(os.path.join(this_file_dir, f'save_figs/eval_topk{fig_suffix}'), dpi=300)